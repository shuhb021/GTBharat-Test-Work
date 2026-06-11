"""
FAR Automation Tool — Excel Parser
Parses client financial statements (Balance Sheet, P&L, Notes to Accounts)
from Excel files using openpyxl.
"""

import re
import logging
import zipfile
import xml.etree.ElementTree as ET
import io
from datetime import datetime
from openpyxl import load_workbook
import os

_WB_CACHE = {}

def get_cached_workbook(filepath):
    if not os.path.exists(filepath):
        raise FileNotFoundError(filepath)
    if filepath not in _WB_CACHE:
        logging.getLogger(__name__).info("Loading workbook into cache (read_only=True): %s", filepath)
        try:
            # Handle .xlsx files
            _WB_CACHE[filepath] = load_workbook(filepath, data_only=True, read_only=True)
            return _WB_CACHE[filepath]
        except Exception as e:
            error_msg = str(e)
            # If it's an xlsb/binary Excel file, attempt conversion via LibreOffice first,
            # then pyxlsb2 fallback if conversion isn't available.
            if '.xlsb' in filepath.lower() or 'binary' in error_msg.lower():
                # Try LibreOffice headless conversion to .xlsx
                try:
                    import shutil, subprocess, tempfile, glob
                    soffice = shutil.which('soffice') or shutil.which('libreoffice')
                    if soffice:
                        tmpdir = tempfile.mkdtemp(prefix='far_xlsb_conv_')
                        try:
                            cmd = [soffice, '--headless', '--convert-to', 'xlsx', filepath, '--outdir', tmpdir]
                            subprocess.run(cmd, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
                            converted = glob.glob(os.path.join(tmpdir, '*.xlsx'))
                            if converted:
                                conv_path = converted[0]
                                logging.getLogger(__name__).info("Converted .xlsb to .xlsx via LibreOffice: %s", conv_path)
                                _WB_CACHE[filepath] = load_workbook(conv_path, data_only=True, read_only=True)
                                return _WB_CACHE[filepath]
                        finally:
                            pass
                except Exception:
                    pass

                # Try PowerShell + Excel COM automation on Windows
                try:
                    if os.name == 'nt':
                        import tempfile, subprocess
                        tmpdir = tempfile.mkdtemp(prefix='far_xlsb_ps_')
                        outpath = os.path.join(tmpdir, os.path.splitext(os.path.basename(filepath))[0] + '.xlsx')
                        # escape single quotes
                        escaped_filepath = filepath.replace("'", "''")
                        escaped_outpath = outpath.replace("'", "''")
                        ps_script = (
                            "$xl = New-Object -ComObject Excel.Application;"
                            "$xl.Visible = $false;"
                            f"$wb = $xl.Workbooks.Open('{escaped_filepath}');"
                            f"$wb.SaveAs('{escaped_outpath}', 51);"
                            "$wb.Close(); $xl.Quit();"
                        )
                        try:
                            subprocess.run(["powershell", "-NoProfile", "-Command", ps_script], check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
                            if os.path.exists(outpath):
                                logging.getLogger(__name__).info("Converted .xlsb to .xlsx via PowerShell/Excel COM: %s", outpath)
                                _WB_CACHE[filepath] = load_workbook(outpath, data_only=True, read_only=True)
                                return _WB_CACHE[filepath]
                        except Exception:
                            pass
                except Exception:
                    pass

                try:
                    try:
                        # prefer pyxlsb2 package when available
                        import pyxlsb2 as _pyxlsb_mod
                        # Monkeypatch function_names to provide safe default for unknown ids
                        try:
                            from collections import defaultdict
                            if hasattr(_pyxlsb_mod, 'ptgs') and hasattr(_pyxlsb_mod.ptgs, 'function_names'):
                                _pyxlsb_mod.ptgs.function_names = defaultdict(lambda: [''], _pyxlsb_mod.ptgs.function_names)
                        except Exception:
                            pass
                        xlsb_open = _pyxlsb_mod.open_workbook
                    except Exception:
                        try:
                            from pyxlsb import open_workbook as xlsb_open
                        except Exception:
                            raise

                    class _XlsbSheet:
                        def __init__(self, sheet):
                            self._sheet = sheet
                            # materialize rows
                            self._rows = list(sheet.rows())
                            self.max_row = len(self._rows)
                            self.max_column = 0
                            for r in self._rows:
                                self.max_column = max(self.max_column, len(r))
                            # emulate sheet_state
                            self.sheet_state = 'visible'

                        def iter_rows(self, values_only=False):
                            for r in self._rows:
                                cells = []
                                for v in r:
                                    val = None
                                    try:
                                        val = v.v
                                    except Exception:
                                        try:
                                            val = v.value
                                        except Exception:
                                            val = v
                                    class _Cell:
                                        def __init__(self, value):
                                            self.value = value
                                            # minimal font stub
                                            self.font = type('F', (), {'bold': False})()
                                    cells.append(_Cell(val))
                                yield tuple(cells)

                    class _XlsbWB:
                        def __init__(self, path):
                            self._sheets = {}
                            self.sheetnames = []
                            self._wb = xlsb_open(path)
                            # pyxlsb open_workbook exposes .sheets or .sheet_names
                            try:
                                names = list(self._wb.sheets)
                            except Exception:
                                try:
                                    names = list(self._wb.sheet_names)
                                except Exception:
                                    names = []
                            for name in names:
                                self.sheetnames.append(name)
                                sh = self._wb.get_sheet(name)
                                self._sheets[name] = _XlsbSheet(sh)

                        def __contains__(self, name):
                            return name in self._sheets

                        def __getitem__(self, name):
                            return self._sheets[name]

                    wb_xlsb = _XlsbWB(filepath)
                    _WB_CACHE[filepath] = wb_xlsb
                    logging.getLogger(__name__).info("Loaded .xlsb via pyxlsb fallback: %s", filepath)
                    return _WB_CACHE[filepath]
                except Exception as e2:
                    # If fallback fails, show user-friendly error
                    logging.getLogger(__name__).exception("pyxlsb fallback failed")
                    raise ValueError(
                        f"ERROR: openpyxl cannot read .xlsb (Excel binary) files.\n"
                        f"File: {filepath}\n"
                        f"Please convert this file to .xlsx format first, or ensure 'pyxlsb2' is installed.\n"
                        f"Original error: {error_msg}\nFallback error: {e2}"
                    ) from e2
            else:
                raise ValueError(f"Failed to load workbook {filepath}: {error_msg}") from e
    
    return _WB_CACHE[filepath]

def clear_workbook_cache():
    global _WB_CACHE
    for wb in _WB_CACHE.values():
        wb.close()
    _WB_CACHE.clear()

def get_hidden_rows_cols(filepath, sheet_name):
    """
    Extract hidden row indices and column indices from the underlying XML structures.

    Handles:
    - Rows with hidden='1' or hidden='true' (manually hidden)
    - Rows with ht='0' (zero-height = invisible)
    - Filter-hidden rows: Excel sets hidden=1 on rows excluded by AutoFilter
    - Columns with hidden='1' or width='0'

    Rules applied:
    - HUMAN VISIBILITY RULE (Rule 1, 3, 4):
      Any row/column not visible to a human in Excel is treated as NON-EXISTENT.
    """
    hidden_rows = set()
    hidden_cols = set()
    try:
        with zipfile.ZipFile(filepath, 'r') as z:
            # 1. Parse workbook.xml to map sheet names to relationship ID (rId)
            wb_xml = z.read('xl/workbook.xml')
            wb_root = ET.fromstring(wb_xml)
            ns = {'ns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
            r_id = None
            for sheet in wb_root.findall('.//ns:sheet', ns):
                if sheet.get('name') == sheet_name:
                    r_id = sheet.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
                    break
            
            if not r_id:
                return hidden_rows, hidden_cols
                
            # 2. Parse workbook.xml.rels to map rId to target sheet file path
            rels_xml = z.read('xl/_rels/workbook.xml.rels')
            rels_root = ET.fromstring(rels_xml)
            target = None
            for rel in rels_root.findall('.//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
                if rel.get('Id') == r_id:
                    target = rel.get('Target')
                    break
                    
            if not target:
                return hidden_rows, hidden_cols
                
            # Normalize target path to start with 'xl/' if it doesn't
            target_clean = target.lstrip('/')
            if not target_clean.startswith('xl/'):
                target_path = 'xl/' + target_clean
            else:
                target_path = target_clean
                
            # 3. Read sheet XML and extract hidden elements
            #    NOTE: Excel sets hidden=1 on ALL filter-excluded rows automatically.
            #          This single check covers manual-hide, zero-height AND filter-hidden rows.
            sheet_xml = z.read(target_path)
            for event, elem in ET.iterparse(io.BytesIO(sheet_xml), events=('start',)):
                tag = elem.tag.split('}')[-1]
                if tag == 'col':
                    is_hidden = elem.get('hidden') in ('1', 'true')
                    width_val = elem.get('width')
                    if width_val is not None:
                        try:
                            if float(width_val) == 0:
                                is_hidden = True
                        except ValueError:
                            pass
                    if is_hidden:
                        c_min = int(elem.get('min'))
                        c_max = int(elem.get('max'))
                        for c in range(c_min, c_max + 1):
                            hidden_cols.add(c)
                elif tag == 'row':
                    # hidden=1 covers: manually hidden rows, zero-height rows,
                    # AND AutoFilter-excluded rows (Excel sets hidden=1 for all three)
                    is_hidden = elem.get('hidden') in ('1', 'true')
                    ht_val = elem.get('ht')
                    if ht_val is not None:
                        try:
                            if float(ht_val) == 0:
                                is_hidden = True
                        except ValueError:
                            pass
                    if is_hidden:
                        r_val = elem.get('r')
                        if r_val is not None:
                            try:
                                hidden_rows.add(int(r_val))
                            except (ValueError, TypeError):
                                pass
                elem.clear()
    except Exception as e:
        import logging
        logging.getLogger(__name__).warning("Error getting hidden rows/cols for sheet '%s': %s", sheet_name, e)
    return hidden_rows, hidden_cols

class SheetWrapper:
    """Wraps a read-only worksheet to support ws.cell(row, col) syntax using an in-memory 2D array."""
    def __init__(self, ws, hidden_rows=None, hidden_cols=None):
        self.cells = {}
        self.max_row = ws.max_row
        self.max_column = ws.max_column
        self.hidden_rows = hidden_rows if hidden_rows is not None else set()
        self.hidden_cols = hidden_cols if hidden_cols is not None else set()
        
        # Load values using iter_rows so we can extract font bold status.
        # This is required for note heading detection and UI formatting.
        for r_idx, row in enumerate(ws.iter_rows(values_only=False), 1):
            if r_idx in self.hidden_rows:
                continue
            for c_idx, cell in enumerate(row, 1):
                if c_idx in self.hidden_cols:
                    continue
                if cell.value is not None:
                    try:
                        is_bold = bool(cell.font and cell.font.bold)
                    except Exception:
                        is_bold = False
                    self.cells[(r_idx, c_idx)] = (cell.value, is_bold)

    def cell(self, row, column):
        data = self.cells.get((row, column))
        class CellStub:
            def __init__(self, d):
                if d is not None:
                    self.value = d[0]
                    self.font = type('FontStub', (), {'bold': d[1]})()
                else:
                    self.value = None
                    self.font = None
        return CellStub(data)


logger = logging.getLogger(__name__)


def _is_bold(cell):
    """Check if a cell's font is bold."""
    try:
        return cell.font and cell.font.bold
    except Exception:
        return False


def _get_value(cell):
    """
    Get numeric value from cell. Returns:
      - float        : explicit number (including 0.0 for explicit zero)
      - '-' (string) : cell contains a dash / nil marker ('-', '—', 'nil', 'n/a')
                       → display as '-' in output
      - None         : cell is truly blank / empty
                       → display as '-NA-' in output

    Callers must handle all three cases.
    """
    v = cell.value
    if v is None:
        return None  # truly blank → -NA-
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        stripped = v.strip()
        if stripped == '':
            return None  # empty string → -NA-
        # Dash / nil markers → sentinel '-'
        if stripped in ('-', '--', '—', '–', 'nil', 'n/a', 'na', 'N/A', 'NA', 'Nil'):
            return '-'
        # Try to parse as number after removing commas
        cleaned = stripped.replace(',', '').replace(' ', '')
        if cleaned.startswith('(') and cleaned.endswith(')'):
            cleaned = '-' + cleaned[1:-1]
        try:
            return float(cleaned)
        except (ValueError, TypeError):
            return None  # non-numeric, non-dash text → -NA-
    return None


def _get_text(cell):
    """Get text value from cell, return empty string if None."""
    v = cell.value
    if v is None:
        return ''
    return str(v).strip()


def _extract_client_name(ws):
    """Heuristic to extract client/company name from the top of the sheet.

    Strategy:
    - Scan the first 6 rows and first 6 columns.
    - Prefer bold cells and non-empty strings that are not obvious headers
      like 'Balance Sheet', 'Profit & Loss', 'STATEMENT', 'AS AT', etc.
    - Return the longest candidate (after preferring bold ones).
    """
    candidates = []
    reject_words = ('balance', 'sheet', 'profit', 'loss', 'statement', 'as at', 'for the', 'audited')
    for r in range(1, 7):
        for c in range(1, 7):
            try:
                cell = ws.cell(row=r, column=c)
            except Exception:
                continue
            txt = _get_text(cell)
            if not txt:
                continue
            low = txt.lower()
            if any(w in low for w in reject_words):
                continue
            is_bold = False
            try:
                is_bold = bool(cell.font and getattr(cell.font, 'bold', False))
            except Exception:
                is_bold = False
            candidates.append((txt, is_bold))

    if not candidates:
        return ''

    # Prefer bold candidates; otherwise pick longest string
    bolds = [t for t, b in candidates if b]
    if bolds:
        return max(bolds, key=lambda s: len(s)).strip()
    return max([t for t, b in candidates], key=lambda s: len(s)).strip()


def _find_date_columns(ws, search_rows=range(5, 12), default_cy_year=2025, default_py_year=2024):
    """
    Find CY and PY date header columns by scanning for datetime(YYYY, 3, 31) values or year text.
    Returns (cy_col, py_col, cy_year, py_year, header_row).
    """
    date_cells = []
    hidden_rows = getattr(ws, 'hidden_rows', set())
    hidden_cols = getattr(ws, 'hidden_cols', set())
    for row_idx in search_rows:
        if row_idx in hidden_rows:
            continue
        for col_idx in range(1, ws.max_column + 1):
            if not os.path.exists(filepath):
                raise FileNotFoundError(filepath)
            if filepath not in _WB_CACHE:
                logging.getLogger(__name__).info("Loading workbook into cache (read_only=True): %s", filepath)

                # Prefer native openpyxl for .xlsx/.xlsm files
                if not filepath.lower().endswith('.xlsb'):
                    try:
                        _WB_CACHE[filepath] = load_workbook(filepath, data_only=True, read_only=True)
                        return _WB_CACHE[filepath]
                    except Exception as e:
                        raise ValueError(f"Failed to load workbook {filepath}: {e}") from e

                # For .xlsb files try two fallbacks: LibreOffice conversion, then pyxlsb parsing
                # 1) LibreOffice conversion (soffice) to temporary .xlsx
                try:
                    import shutil, subprocess, tempfile
                    soffice = shutil.which('soffice') or shutil.which('libreoffice')
                    if soffice:
                        tmpdir = tempfile.mkdtemp(prefix='far_xlsb_')
                        try:
                            cmd = [soffice, '--headless', '--convert-to', 'xlsx', filepath, '--outdir', tmpdir]
                            subprocess.run(cmd, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
                            # Find converted file in tmpdir
                            import glob
                            converted = glob.glob(os.path.join(tmpdir, '*.xlsx'))
                            if converted:
                                conv_path = converted[0]
                                logging.getLogger(__name__).info("Converted .xlsb to .xlsx via LibreOffice: %s", conv_path)
                                _WB_CACHE[filepath] = load_workbook(conv_path, data_only=True, read_only=True)
                                return _WB_CACHE[filepath]
                        finally:
                            try:
                                # Do not remove immediately; rely on OS temp cleanup
                                pass
                            except Exception:
                                pass
                except Exception:
                    # Fall through to pyxlsb attempt
                    pass

                # 1b) On Windows, try PowerShell + Excel COM automation to convert .xlsb -> .xlsx
                try:
                    if os.name == 'nt':
                        import tempfile, subprocess
                        tmpdir = tempfile.mkdtemp(prefix='far_xlsb_ps_')
                        outpath = os.path.join(tmpdir, os.path.splitext(os.path.basename(filepath))[0] + '.xlsx')
                        # Escape single quotes for PowerShell single-quoted strings
                        escaped_filepath = filepath.replace("'", "''")
                        escaped_outpath = outpath.replace("'", "''")
                        ps_script = (
                            "$xl = New-Object -ComObject Excel.Application;"
                            "$xl.Visible = $false;"
                            f"$wb = $xl.Workbooks.Open('{escaped_filepath}');"
                            f"$wb.SaveAs('{escaped_outpath}', 51);"
                            "$wb.Close(); $xl.Quit();"
                        )
                        try:
                            subprocess.run(["powershell", "-NoProfile", "-Command", ps_script], check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
                            if os.path.exists(outpath):
                                logging.getLogger(__name__).info("Converted .xlsb to .xlsx via PowerShell/Excel COM: %s", outpath)
                                _WB_CACHE[filepath] = load_workbook(outpath, data_only=True, read_only=True)
                                return _WB_CACHE[filepath]
                        except Exception:
                            pass
                except Exception:
                    pass

                # 2) pyxlsb parsing fallback: build a minimal workbook-like wrapper
                try:
                    try:
                        from pyxlsb import open_workbook as xlsb_open
                    except Exception:
                        # Some installs use 'pyxlsb2' package name with same API
                        from pyxlsb import open_workbook as xlsb_open

                    class _XlsbSheet:
                        def __init__(self, sheet):
                            self._sheet = sheet
                            # Build 2D grid
                            rows = list(sheet.rows())
                            self._rows = rows
                            self.max_row = len(rows)
                            self.max_column = 0
                            for r in rows:
                                # r is iterable of Cell objects or tuples
                                self.max_column = max(self.max_column, len(r))

                        def iter_rows(self, values_only=False):
                            for r in self._rows:
                                cells = []
                                for v in r:
                                    # v may be a Cell object or raw value
                                    val = None
                                    try:
                                        val = v.v
                                    except Exception:
                                        val = v
                                    class _Cell:
                                        def __init__(self, value):
                                            self.value = value
                                            class _F: pass
                                            self.font = type('F', (), {'bold': False})()
                                    cells.append(_Cell(val))
                                # Pad to max_column
                                yield tuple(cells)

                    class _XlsbWB:
                        def __init__(self, path):
                            self._sheets = {}
                            self.sheetnames = []
                            self._wb = xlsb_open(path)
                            for name in self._wb.sheets:
                                self.sheetnames.append(name)
                                sh = self._wb.get_sheet(name)
                                self._sheets[name] = _XlsbSheet(sh)

                        def __getitem__(self, name):
                            return self._sheets[name]

                    wb_xlsb = _XlsbWB(filepath)
                    _WB_CACHE[filepath] = wb_xlsb
                    logging.getLogger(__name__).info("Loaded .xlsb via pyxlsb fallback: %s", filepath)
                    return _WB_CACHE[filepath]
                except Exception as e:
                    logging.getLogger(__name__).error("Failed to parse .xlsb via pyxlsb: %s", e)
                    raise ValueError(
                        f"ERROR: openpyxl cannot read .xlsb (Excel binary) files.\n"
                        f"File: {filepath}\n"
                        f"Please convert this file to .xlsx format first, or install LibreOffice for automatic conversion.\n"
                        f"Original error: {e}"
                    ) from e

            return _WB_CACHE[filepath]
    hidden_cols = getattr(ws, 'hidden_cols', set())
    
    # 1. Scan header row and adjacent rows
    for r in range(max(1, header_row - 2), min(ws.max_row + 1, header_row + 2)):
        if r in hidden_rows:
            continue
        for c in range(1, ws.max_column + 1):
            if c in hidden_cols:
                continue
            val = str(ws.cell(row=r, column=c).value or '').lower().strip()
            if not found_part and ('particular' in val or 'description' in val):
                part_col = c
                found_part = True
            elif not found_note and ('note' in val or 'sch' in val):
                note_col = c
                found_note = True
                
    # 2. Fallback heuristic: check which column has the most text in the data rows (usually particulars)
    if not found_part:
        col_lengths = {1: 0, 2: 0, 3: 0, 4: 0}
        for r in range(header_row + 1, min(ws.max_row + 1, header_row + 15)):
            if r in hidden_rows:
                continue
            for c in range(1, 5):
                if c in hidden_cols:
                    continue
                val = str(ws.cell(row=r, column=c).value or '').strip()
                # Don't count numbers/dates
                if not re.match(r'^[\d\.,\(\)\-]+$', val):
                    col_lengths[c] += len(val)
                    
        best_col = max(col_lengths.items(), key=lambda x: x[1])
        if best_col[1] > 20:
            part_col = best_col[0]
            logger.info("Heuristic determined Particulars is in column %d (length sum=%d)", part_col, best_col[1])
            
    return part_col, note_col


def _append_adjustments(ws, row_idx, part_text, part_col, cy_col, py_col):
    """Scan columns for manual/round-off adjustments and append to particulars text."""
    for c in range(1, ws.max_column + 1):
        if c in (part_col, cy_col, py_col):
            continue
        val = _get_text(ws.cell(row=row_idx, column=c))
        if val and any(x in val.lower() for x in ['manual adj', 'round off adj', 'rounding off adj', 'matching']):
            clean_val = val.strip()
            if clean_val.lower() not in part_text.lower():
                return f"{part_text} ({clean_val})"
    return part_text


def parse_balance_sheet(filepath):
    """
    Parse Balance Sheet from client Excel file.

    Expected structure (from sample):
    - Row 1: Client name
    - Row 3: "Balance Sheet as at 31 March YYYY"
    - Row 7: Headers with datetime(YYYY, 3, 31) in cols G (CY) and I (PY)
    - Col A: Particulars
    - Col C: Note numbers
    - Data from "Assets" row to "Total equity and liabilities" row

    Returns:
        dict with keys:
            'data': list of row dicts
            'cy_year': int
            'py_year': int
            'client_name': str
    """
    wb = get_cached_workbook(filepath)

    # Try to find 'BS' sheet or first sheet that is visible
    target_sheet_name = None
    for name in ['BS', 'Balance Sheet', 'BalanceSheet']:
        if name in wb.sheetnames and wb[name].sheet_state == 'visible':
            target_sheet_name = name
            break
            
    if target_sheet_name is None:
        for name in wb.sheetnames:
            if wb[name].sheet_state == 'visible':
                target_sheet_name = name
                break
                
    if target_sheet_name is None:
        raise ValueError("No visible worksheets found in the workbook.")

    ws_obj = wb[target_sheet_name]
    hidden_rows, hidden_cols = get_hidden_rows_cols(filepath, target_sheet_name)
    ws = SheetWrapper(ws_obj, hidden_rows=hidden_rows, hidden_cols=hidden_cols)

    # Get client name — enhanced scan across first 5 rows
    client_name = _extract_client_name(ws)

    # Find date columns
    cy_col, py_col, cy_year, py_year, header_row = _find_date_columns(ws)
    if cy_col is None:
        raise ValueError("Could not find year headers (e.g. '31 March 2025') in the Balance Sheet.\n"
                         "Please load your actual client financial statement, not the blank template.")

    # Determine particulars and notes columns dynamically
    part_col, note_col = _find_part_note_columns(ws, header_row, default_part=1, default_note=3)
    logger.info("BS columns detected: Particulars=col %d, Notes=col %d", part_col, note_col)

    # Parse data rows
    data = []
    notes = []
    signatures = []
    footers = []
    is_table_closed = False
    start_row = header_row + 1

    for row_idx in range(start_row, ws.max_row + 1):
        if row_idx in ws.hidden_rows:
            continue

        part_text = _get_text(ws.cell(row=row_idx, column=part_col))
        cy_val = _get_value(ws.cell(row=row_idx, column=cy_col))
        py_val = _get_value(ws.cell(row=row_idx, column=py_col))
        is_bold = _is_bold(ws.cell(row=row_idx, column=part_col))
        note_text = _get_text(ws.cell(row=row_idx, column=note_col))

        # Skip completely empty rows
        if not part_text and cy_val is None and py_val is None:
            continue

        row_class = _classify_row(ws, row_idx, part_text, cy_val, py_val, is_bold)

        if is_table_closed:
            row_cells = _extract_row_cells(ws, row_idx)
            if any(c != '' and c != '-' for c in row_cells):
                if row_class == 'SIGNATURE':
                    signatures.append(row_cells)
                elif row_class == 'NOTE' or _is_narrative_text(part_text) or len(str(part_text)) > 40:
                    notes.append(row_cells)
                else:
                    footers.append(row_cells)
            continue

        # If not closed:
        if row_class == 'SIGNATURE':
            is_table_closed = True
            signatures.append(_extract_row_cells(ws, row_idx))
            continue

        if row_class == 'NOTE':
            notes.append(_extract_row_cells(ws, row_idx))
            continue

        if row_class == 'FOOTER':
            footers.append(_extract_row_cells(ws, row_idx))
            continue

        # FINANCIAL_DATA
        part_text = _append_adjustments(ws, row_idx, part_text, part_col, cy_col, py_col)
        row_data = {
            'particulars': part_text,
            'note': note_text,
            'cy': cy_val if cy_val is not None else None,
            'py': py_val if py_val is not None else None,
            'is_bold': is_bold,
            'row_num': row_idx
        }
        data.append(row_data)

        # Table Continuation Check
        part_lower = part_text.lower().strip()
        
        # Absolute immediate termination for the Balance Sheet
        if 'total equity and liabilities' in part_lower:
            is_table_closed = True
            
        elif "total" in part_lower:
            close_table = True
            for next_row_idx in range(row_idx + 1, min(row_idx + 11, ws.max_row + 1)):
                if next_row_idx in ws.hidden_rows:
                    continue
                next_part = _get_text(ws.cell(row=next_row_idx, column=part_col))
                if not next_part or next_part.strip() == '':
                    continue
                next_cy = _get_value(ws.cell(row=next_row_idx, column=cy_col))
                next_py = _get_value(ws.cell(row=next_row_idx, column=py_col))
                next_bold = _is_bold(ws.cell(row=next_row_idx, column=part_col))
                next_class = _classify_row(ws, next_row_idx, next_part, next_cy, next_py, next_bold)
                if next_class == 'FINANCIAL_DATA':
                    if next_cy is not None or next_py is not None:
                        close_table = False
                        break
            if close_table:
                is_table_closed = True

    logger.info("Parsed BS: %d rows, CY=%d, PY=%d, notes=%d, signatures=%d, footers=%d",
                len(data), cy_year, py_year, len(notes), len(signatures), len(footers))

    return {
        'data': data,
        'notes': notes,
        'signatures': signatures,
        'footers': footers,
        'cy_year': cy_year,
        'py_year': py_year,
        'client_name': client_name,
        'cy_col_letter': _col_letter(cy_col),
        'py_col_letter': _col_letter(py_col)
    }


def parse_profit_loss(filepath):
    """
    Parse Profit & Loss statement from client Excel file.

    Expected structure (from sample):
    - Col A: Particulars
    - Col D: Note numbers
    - Col E: CY values, Col G: PY values (row 7 has dates)
    - Data from "Revenue from operations" to end

    Returns:
        dict with keys: 'data', 'cy_year', 'py_year', 'client_name'
    """
    wb = get_cached_workbook(filepath)

    # Try to find 'PL' or 'P&L' sheet that is visible
    target_sheet_name = None
    for name in ['PL', 'P&L', 'Profit & Loss', 'Profit and Loss', 'ProfitLoss']:
        if name in wb.sheetnames and wb[name].sheet_state == 'visible':
            target_sheet_name = name
            break
            
    if target_sheet_name is None:
        # Check first visible sheet that is not BS or cover page
        for name in wb.sheetnames:
            name_lower = name.lower()
            if wb[name].sheet_state == 'visible' and not any(x in name_lower for x in ['bs', 'balance sheet', 'cover page', 'dashboard']):
                target_sheet_name = name
                break

    if target_sheet_name is None:
        # Fallback to first visible sheet
        for name in wb.sheetnames:
            if wb[name].sheet_state == 'visible':
                target_sheet_name = name
                break

    if target_sheet_name is None:
        raise ValueError("No visible worksheets found in the workbook.")

    ws_obj = wb[target_sheet_name]
    hidden_rows, hidden_cols = get_hidden_rows_cols(filepath, target_sheet_name)
    ws = SheetWrapper(ws_obj, hidden_rows=hidden_rows, hidden_cols=hidden_cols)

    client_name = _extract_client_name(ws)

    # Find date columns
    cy_col, py_col, cy_year, py_year, header_row = _find_date_columns(ws)
    if cy_col is None:
        raise ValueError("Could not find year headers (e.g. '31 March 2025') in the P&L statement.\n"
                         "Please load your actual client financial statement, not the blank template.")

    # Determine particulars and notes columns dynamically
    part_col, note_col = _find_part_note_columns(ws, header_row, default_part=1, default_note=4)
    logger.info("PL columns detected: Particulars=col %d, Notes=col %d", part_col, note_col)

    data = []
    notes = []
    signatures = []
    footers = []
    is_table_closed = False
    start_row = header_row + 1

    for row_idx in range(start_row, ws.max_row + 1):
        if row_idx in ws.hidden_rows:
            continue

        part_text = _get_text(ws.cell(row=row_idx, column=part_col))
        cy_val = _get_value(ws.cell(row=row_idx, column=cy_col))
        py_val = _get_value(ws.cell(row=row_idx, column=py_col))
        is_bold = _is_bold(ws.cell(row=row_idx, column=part_col))
        note_text = _get_text(ws.cell(row=row_idx, column=note_col))

        # Skip completely empty rows
        if not part_text and cy_val is None and py_val is None:
            continue

        row_class = _classify_row(ws, row_idx, part_text, cy_val, py_val, is_bold)

        if is_table_closed:
            row_cells = _extract_row_cells(ws, row_idx)
            if any(c != '' and c != '-' for c in row_cells):
                if row_class == 'SIGNATURE':
                    signatures.append(row_cells)
                elif row_class == 'NOTE' or _is_narrative_text(part_text) or len(str(part_text)) > 40:
                    notes.append(row_cells)
                else:
                    footers.append(row_cells)
            continue

        # If not closed:
        if row_class == 'SIGNATURE':
            is_table_closed = True
            signatures.append(_extract_row_cells(ws, row_idx))
            continue

        if row_class == 'NOTE':
            notes.append(_extract_row_cells(ws, row_idx))
            continue

        if row_class == 'FOOTER':
            footers.append(_extract_row_cells(ws, row_idx))
            continue

        # FINANCIAL_DATA
        part_text = _append_adjustments(ws, row_idx, part_text, part_col, cy_col, py_col)
        row_data = {
            'particulars': part_text,
            'note': note_text,
            'cy': cy_val if cy_val is not None else None,
            'py': py_val if py_val is not None else None,
            'is_bold': is_bold,
            'row_num': row_idx
        }
        data.append(row_data)

        # Table Continuation Check
        part_lower = part_text.lower().strip()
        
        # Absolute immediate termination for the Profit & Loss statement
        if any(term in part_lower for term in ['total comprehensive income', 'profit / (loss) for the year', 'profit/(loss) for the year']):
            is_table_closed = True
            
        elif "total" in part_lower or "profit" in part_lower or "loss" in part_lower:
            close_table = True
            for next_row_idx in range(row_idx + 1, min(row_idx + 11, ws.max_row + 1)):
                if next_row_idx in ws.hidden_rows:
                    continue
                next_part = _get_text(ws.cell(row=next_row_idx, column=part_col))
                if not next_part or next_part.strip() == '':
                    continue
                next_cy = _get_value(ws.cell(row=next_row_idx, column=cy_col))
                next_py = _get_value(ws.cell(row=next_row_idx, column=py_col))
                next_bold = _is_bold(ws.cell(row=next_row_idx, column=part_col))
                next_class = _classify_row(ws, next_row_idx, next_part, next_cy, next_py, next_bold)
                if next_class == 'FINANCIAL_DATA':
                    if next_cy is not None or next_py is not None:
                        close_table = False
                        break
            if close_table:
                is_table_closed = True

    logger.info("Parsed PL: %d rows, CY=%d, PY=%d, notes=%d, signatures=%d, footers=%d",
                len(data), cy_year, py_year, len(notes), len(signatures), len(footers))

    return {
        'data': data,
        'notes': notes,
        'signatures': signatures,
        'footers': footers,
        'cy_year': cy_year,
        'py_year': py_year,
        'client_name': client_name,
        'cy_col_letter': _col_letter(cy_col),
        'py_col_letter': _col_letter(py_col)
    }


def parse_notes(filepath, cy_year=2025, py_year=2024):
    """
    Parse Notes to Accounts from client Excel file.
    Handles multiple sheets (3-4, 5-9, 10-17, 18-26) each containing
    multiple notes with their own headers.

    Returns:
        dict mapping sheet_name → list of note groups:
        {
            '3-4': [
                {
                    'note_num': '3',
                    'note_heading': 'Property and equipment',
                    'data': [...],
                    'is_asset_note': True/False
                },
                ...
            ],
            ...
        }
    """
    wb = get_cached_workbook(filepath)

    # Dynamically find notes sheets instead of hardcoding
    notes_sheet_names = []
    for sheet_name in wb.sheetnames:
        name_lower = sheet_name.lower().strip()
        
        # Skip hidden sheets
        if wb[sheet_name].sheet_state != 'visible':
            continue
            
        # Skip known main sheets
        if any(x in name_lower for x in ['bs', 'balance sheet', 'pl', 'p&l', 'profit']):
            continue
        if name_lower in ['dashboard', 'ratios', 'cover page']:
            continue
            
        # If the sheet name contains digits (e.g. '3-4', '5') or the word 'note'/'sch', it's a notes sheet
        if re.search(r'\d', sheet_name) or 'note' in name_lower or 'sch' in name_lower:
            notes_sheet_names.append(sheet_name)
            
    notes_dict = {}
    sheet_notes_dict = {}
    signatures_dict = {}
    footers_dict = {}

    for sheet_name in notes_sheet_names:
        if sheet_name not in wb.sheetnames:
            logger.debug("Notes sheet '%s' not found in workbook", sheet_name)
            continue

        hidden_rows, hidden_cols = get_hidden_rows_cols(filepath, sheet_name)
        ws = SheetWrapper(wb[sheet_name], hidden_rows=hidden_rows, hidden_cols=hidden_cols)
        notes_groups, sheet_notes, signatures, footers = _parse_notes_sheet(ws, cy_year, py_year)
        
        if notes_groups:
            notes_dict[sheet_name] = notes_groups
        if sheet_notes:
            sheet_notes_dict[sheet_name] = sheet_notes
        if signatures:
            signatures_dict[sheet_name] = signatures
        if footers:
            footers_dict[sheet_name] = footers

    if not notes_dict:
        logger.warning("No notes were parsed. This usually happens if you load a blank template instead of actual client data.")
    else:
        logger.info("Parsed notes from %d sheets, footers from %d sheets", len(notes_dict), len(footers_dict))
    return {
        'notes': notes_dict,
        'sheet_notes': sheet_notes_dict,
        'signatures': signatures_dict,
        'footers': footers_dict
    }


def _is_note_heading(ws, row_idx):
    """
    Check if a row starts a new note group.
    Returns (note_num, note_heading, heading_col) or (None, None, None).
    """
    if row_idx in getattr(ws, 'hidden_rows', set()):
        return None, None, None

    hidden_cols = getattr(ws, 'hidden_cols', set())

    # 1. First, check columns A, B, C for "Note X:" pattern
    for col_idx in (1, 2, 3):
        if col_idx in hidden_cols:
            continue
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.value is not None:
            val_str = str(cell.value).strip()
            # Match "Note 3", "Note 3(a)", "Note 4: ...", "Note 4.1"
            m = re.match(r'^Note\s*(\d+(?:\.\d+)?(?:[A-Za-z]|\([a-z]\))?)\s*[:\-]?\s*(.*)$', val_str, re.IGNORECASE)
            if m:
                note_num = m.group(1)
                note_heading = m.group(2).strip()
                heading_col_idx = col_idx
                if not note_heading:
                    # Look in the next column for the heading
                    note_heading = _get_text(ws.cell(row=row_idx, column=col_idx + 1))
                    heading_col_idx = col_idx + 1
                heading_clean = note_heading.lower().strip()
                if heading_clean and heading_clean not in ['particulars', 'total', 'amount', 'rs', 'in lakhs', 'in lacs', 'in millions', 'in thousands', 'rs.', 'rupees']:
                    return note_num, note_heading, heading_col_idx

    # 2. Check for standalone note numbers in Column A or B
    a_cell = ws.cell(row=row_idx, column=1)
    b_cell = ws.cell(row=row_idx, column=2)
    c_cell = ws.cell(row=row_idx, column=3)

    def is_valid_note_num(val_str):
        m = re.match(r'^(\d+)(?:\.\d+)?(?:[A-Za-z]|\([a-z]\))?$', val_str)
        if m:
            try:
                num_val = int(m.group(1))
                if num_val > 100:
                    return False
            except ValueError:
                pass
            return True
        return False

    # Check if Column A has note number and Column B has heading
    if 1 not in hidden_cols and 2 not in hidden_cols and a_cell.value is not None:
        a_val = str(a_cell.value).strip()
        if is_valid_note_num(a_val):
            heading = _get_text(b_cell)
            heading_clean = heading.lower().strip()
            if heading and heading_clean not in ['particulars', 'total', 'amount', 'rs', 'in lakhs', 'in lacs', 'in millions', 'in thousands', 'rs.', 'rupees']:
                return a_val, heading, 2

    # Check if Column B has note number and Column C has heading (very common)
    if 2 not in hidden_cols and 3 not in hidden_cols and b_cell.value is not None:
        b_val = str(b_cell.value).strip()
        if is_valid_note_num(b_val):
            heading = _get_text(c_cell)
            heading_clean = heading.lower().strip()
            if heading and heading_clean not in ['particulars', 'total', 'amount', 'rs', 'in lakhs', 'in lacs', 'in millions', 'in thousands', 'rs.', 'rupees']:
                # For standalone numbers in column B, require that either B or C is bold to avoid false positives
                if _is_bold(b_cell) or _is_bold(c_cell):
                    return b_val, heading, 3

    # 3. Fallback for known headings in column B or C (must be bold to avoid false positives)
    for col_idx, note_val in [(2, '3'), (3, '3')]:
        if col_idx in hidden_cols:
            continue
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.value is not None and _is_bold(cell):
            val_str = str(cell.value).strip()
            lower_val = val_str.lower()
            if any(kw in lower_val for kw in ['property', 'plant', 'equipment', 'intangible', 'right of use', 'right-of-use', 'rou', 'ppe']):
                if 'intangible' in lower_val:
                    note_num = '3(b)'
                elif 'right' in lower_val or 'rou' in lower_val:
                    note_num = '3A'
                else:
                    note_num = '3(a)'
                return note_num, val_str, col_idx

    return None, None, None


def _parse_notes_sheet(ws, cy_year=2025, py_year=2024):
    """Parse a single notes sheet, extracting all note groups, sheet-level notes, signatures, and footers."""
    notes_groups = []
    sheet_notes = []
    signatures = []
    footers = []

    hidden_rows = getattr(ws, 'hidden_rows', set())
    hidden_cols = getattr(ws, 'hidden_cols', set())

    row = 1
    while row <= ws.max_row:
        if row in hidden_rows:
            row += 1
            continue

        note_num, note_heading, heading_col = _is_note_heading(ws, row)

        if note_num is None:
            # If not a note heading, classify the row to see if it is a signature, narrative note, or footer
            part_text = ""
            is_bold = False
            for col in range(1, ws.max_column + 1):
                txt = _get_text(ws.cell(row=row, column=col))
                if txt and not re.match(r'^[\d\.,\(\)\-\s]+$', txt):
                    part_text = txt
                    is_bold = _is_bold(ws.cell(row=row, column=col))
                    break
            cy_val = None
            py_val = None
            row_class = _classify_row(ws, row, part_text, cy_val, py_val, is_bold)
            row_cells = _extract_row_cells(ws, row)
            if any(c != '' and c != '-' for c in row_cells):
                if row_class == 'SIGNATURE':
                    signatures.append(row_cells)
                elif row_class == 'NOTE':
                    sheet_notes.append(row_cells)
                elif row_class == 'FOOTER':
                    footers.append(row_cells)
            row += 1
            continue

        # Found a note heading. Now find the date headers below it.
        cy_col, py_col, cy_yr, py_yr, date_row = _find_date_columns(
            ws, search_rows=range(row, min(row + 10, ws.max_row + 1)),
            default_cy_year=cy_year, default_py_year=py_year)

        # Detect if this is an asset note (PPE)
        heading_clean = note_heading.lower().replace('-', ' ')
        is_asset_note = any(kw in heading_clean for kw in
                           ['property', 'plant', 'equipment', 'intangible',
                            'right of use', 'rou', 'ppe'])

        if cy_col is None:
            if is_asset_note:
                # For PPE schedules, find the "Total" column
                found_total = False
                for r in range(row, min(row + 10, ws.max_row + 1)):
                    if r in hidden_rows:
                        continue
                    for c in range(1, ws.max_column + 1):
                        if c in hidden_cols:
                            continue
                        val = _get_text(ws.cell(row=r, column=c)).lower()
                        if val == 'total':
                            cy_col = c
                            py_col = c
                            cy_yr = cy_year
                            py_yr = py_year
                            date_row = r
                            logger.info("Note %s: Found 'Total' column at col %d for Asset note", note_num, c)
                            found_total = True
                            break
                    if found_total:
                        break

            if cy_col is None:
                # Fallback to default columns C (3) and D (4)
                cy_col = 3
                py_col = 4
                if heading_col is not None and heading_col >= cy_col:
                    cy_col = heading_col + 1
                    py_col = heading_col + 2
                cy_yr = cy_year
                py_yr = py_year
                date_row = row + 1
                logger.info("Note %s: Could not find date columns — falling back to columns %d and %d", note_num, cy_col, py_col)
        else:
            cy_year = cy_yr
            py_year = py_yr

        # Find the best column for particulars in the current note dynamically
        part_col = heading_col if heading_col is not None else 2
        
        note_end_row = ws.max_row
        for next_r in range(row + 1, ws.max_row + 1):
            if next_r in hidden_rows:
                continue
            n_num, _, _ = _is_note_heading(ws, next_r)
            if n_num is not None:
                note_end_row = next_r - 1
                break
                
        best_part_col = None
        data_start = date_row + 1 if date_row else row + 3
        for r in range(row, min(data_start + 3, note_end_row + 1)):
            if r in hidden_rows:
                continue
            for c in range(1, ws.max_column + 1):
                if c in hidden_cols or c == cy_col or c == py_col:
                    continue
                val = str(ws.cell(row=r, column=c).value or '').lower().strip()
                if 'particular' in val or 'description' in val or 'detail' in val:
                    best_part_col = c
                    break
            if best_part_col is not None:
                break
                
        if best_part_col is None:
            col_lengths = {}
            for r in range(data_start, min(data_start + 15, note_end_row + 1)):
                if r in hidden_rows:
                    continue
                for c in range(1, min(10, ws.max_column + 1)):
                    if c in hidden_cols or c == cy_col or c == py_col:
                        continue
                    val = str(ws.cell(row=r, column=c).value or '').strip()
                    if val and not re.match(r'^[\d\.,\(\)\-\s%]+$', val):
                        if val.lower() not in ['note', 'no.', 'sch', 'sch.', 'notes', 'particulars', 'description', 'amount', 'rs', 'rupees']:
                            col_lengths[c] = col_lengths.get(c, 0) + len(val)
            if col_lengths:
                best_col = max(col_lengths.items(), key=lambda x: x[1])
                if best_col[1] > 0:
                    best_part_col = best_col[0]
                    
        if best_part_col is not None:
            part_col = best_part_col

        data_rows = []

        is_note_table_closed = False

        data_row = data_start
        while data_row <= ws.max_row:
            if data_row in hidden_rows:
                data_row += 1
                continue

            # Check if we hit the next note heading
            next_num, next_heading, next_col = _is_note_heading(ws, data_row)
            if next_num is not None:
                break  # Next note starts here

            part_text = _get_text(ws.cell(row=data_row, column=part_col))
            cy_val = _get_value(ws.cell(row=data_row, column=cy_col))
            py_val = _get_value(ws.cell(row=data_row, column=py_col))
            is_bold_row = _is_bold(ws.cell(row=data_row, column=part_col))

            if not part_text and cy_val is None and py_val is None:
                data_row += 1
                continue

            row_class = _classify_row(ws, data_row, part_text, cy_val, py_val, is_bold_row)

            if is_note_table_closed:
                row_cells = _extract_row_cells(ws, data_row)
                if any(c != '' and c != '-' for c in row_cells):
                    if row_class == 'SIGNATURE':
                        signatures.append(row_cells)
                    elif row_class == 'NOTE' or _is_narrative_text(part_text) or len(str(part_text)) > 40:
                        sheet_notes.append(row_cells)
                    else:
                        footers.append(row_cells)
                data_row += 1
                row = data_row - 1
                continue

            # If not closed:
            if row_class == 'SIGNATURE':
                is_note_table_closed = True
                signatures.append(_extract_row_cells(ws, data_row))
                data_row += 1
                row = data_row - 1
                continue

            if row_class == 'NOTE':
                sheet_notes.append(_extract_row_cells(ws, data_row))
                data_row += 1
                row = data_row - 1
                continue

            if row_class == 'FOOTER':
                footers.append(_extract_row_cells(ws, data_row))
                data_row += 1
                row = data_row - 1
                continue

            # FINANCIAL_DATA
            part_text = _append_adjustments(ws, data_row, part_text, part_col, cy_col, py_col)
            data_rows.append({
                'particulars': part_text,
                'cy': cy_val if cy_val is not None else None,
                'py': py_val if py_val is not None else None,
                'is_bold': is_bold_row,
                'row_num': data_row
            })

            # Table Continuation Check - Strict Immediate Closure
            part_lower = part_text.lower().strip()
            
            # Absolute immediate termination for specific final totals
            if 'total equity and liabilities' in part_lower:
                is_note_table_closed = True
            elif any(term in part_lower for term in ['total comprehensive income', 'profit / (loss) for the year', 'profit/(loss) for the year']):
                is_note_table_closed = True
            elif any(t in part_lower for t in ['total', 'grand total', 'total amount', 'total value']):
                # Only close if it's an actual total row (has values or is styled as a total)
                if (cy_val is not None and cy_val != '-') or (py_val is not None and py_val != '-') or is_bold_row:
                    is_note_table_closed = True

            data_row += 1
            row = data_row - 1

        notes_groups.append({
            'note_num': note_num,
            'note_heading': note_heading,
            'data': data_rows,
            'is_asset_note': is_asset_note,
            'cy_year': cy_yr,
            'py_year': py_yr
        })

        row += 1

    return notes_groups, sheet_notes, signatures, footers


def _is_footer_row(ws, row_idx):
    # Exact matches (case-insensitive, stripped of non-alphanumeric chars)
    EXACT_ROLES = {
        'partner', 'partners',
        'director', 'directors',
        'managing director',
        'cfo', 'chief financial officer',
        'company secretary',
        'auditor', 'auditors',
        'authorised signatory', 'authorized signatory'
    }

    # High-confidence signature patterns (substring matches)
    HIGH_CONF_SUBSTR = [
        r'\bchartered\s+accountants\b',
        r'\bmembership\s+no\b',
        r'\bmembership\s+number\b',
        r'\bregistration\s+no\b',
        r'\bregistration\s+number\b',
        r'^\s*place\s*:',
        r'^\s*date\s*:',
        r'\budin\b',
        r'\bauthorised\s+signatory\b',
        r'\bauthorized\s+signatory\b',
        r'\bfor\s+and\s+on\s+behalf\s+of\b',
        r'\bas\s+per\s+our\s+report\b',
        r'\bof\s+even\s+date\b',
    ]

    # Matches "For <Company Name>" where company name is letters/spaces and optional LLP/Ltd/Co etc.
    EXCLUDE_WORDS = '(?:the|year|period|note|each|services|raw|other|depreciation|amortisation|amortization|interest|tax|purchase|sale|employee|allowance|provision|doubtful|bad|financial|operating|current|non-current)'
    FOR_PATTERN = re.compile(rf'^for\s+(?!{EXCLUDE_WORDS}\b)[A-Za-z0-9\s&\.,\']+(?:llp|ltd|limited|associates|co|board|directors)?$', re.I)

    for col in range(1, ws.max_column + 1):
        cell_val = ws.cell(row=row_idx, column=col).value
        if cell_val is None:
            continue
        val = str(cell_val).strip()
        if not val:
            continue
        val_lower = val.lower()

        # Check high-confidence substrings
        for p in HIGH_CONF_SUBSTR:
            if re.search(p, val_lower):
                return True

        # Check "For <Company>"
        if FOR_PATTERN.match(val):
            if len(val) < 80:
                return True

        # Check exact roles
        cleaned_val = re.sub(r'[^a-z0-9\s]', '', val_lower).strip()
        if cleaned_val in EXACT_ROLES:
            return True

    return False


def _is_narrative_text(particulars):
    if not particulars:
        return False
    text = particulars.strip()
    text_lower = text.lower()
    
    if text.endswith('.') and not text_lower.endswith(('no.', 'co.', 'llp.', 'ltd.', 'a.m.', 'p.m.', 'i.e.', 'e.g.')):
        return True
        
    narrative_phrases = [
        'represents', 'refers to', 'referred to', 'refer to', 'see note', 'refer note', 
        'during the year', 'previous year', 'current year', 'company has', 'company has not', 
        'amount less than', 'pursuant to', 'disclosed', 'explained', 'clarification', 
        'regrouped', 'reclassified', 'management representation', 'bonus shares', 
        'of even date', 'as per our report', 'carried at', 'measured at', 'accounting policy', 
        'useful lives', 'estimated by', 'depreciation is', 'method of', 'financial statements', 
        'subject to', 'written down', 'residual value', 'useful life', 'in respect of', 
        'valuation of', 'valued by', 'estimated useful', 'carrying amount', 'realisable value', 
        'net book value', 'book value of', 'useful lives of', 'under the', 'agree with', 
        'reconciled to', 'outstanding for a period', 'outstanding for following', 
        'due from directors', 'due from officers', 'due from firms', 'due from private', 
        'held by holding', 'holder of', 'ultimate holding', 'subsidiaries/ associates', 'subsidiary/ associate', 
        'promoter\'s holding', 'promoters\' holding', 'reconciliation of', 'shares outstanding', 
        'beginning of the year', 'end of the year', 'nature and purpose', 're-measurements', 
        'defined benefit', 'continuing operations', 'discontinued operations', 'zero represents', 
        'represents amount', 'less than', 'exceeded its cost', 'overdue'
    ]
    for phrase in narrative_phrases:
        if phrase in text_lower:
            return True
            
    words = re.findall(r'\b[a-zA-Z]{2,}\b', text_lower)
    if not words:
        return False
        
    narrative_verbs = {
        'is', 'are', 'was', 'were', 'have', 'has', 'had', 'been', 'be',
        'does', 'do', 'did', 'should', 'would', 'could', 'will', 'shall', 'may',
        'recommends', 'recommended', 'proposes', 'proposed', 'values', 'valued',
        'defines', 'defined', 'measure', 'measured', 'estimates', 'estimated',
        'expects', 'expected', 'settle', 'settled', 'reclassify', 'reclassified',
        'regroup', 'regrouped', 'opted', 'outstanding', 'disclosed', 'incurred',
        'earned'
    }
    
    narrative_others = {
        'the', 'our', 'their', 'we', 'they', 'it', 'them', 'us', 'who', 'which',
        'whose', 'since', 'because', 'although', 'though', 'about', 'against', 
        'between', 'into', 'through', 'during', 'before', 'after', 'above', 
        'below', 'under', 'over', 'again', 'further', 'then', 'once', 'here', 
        'there', 'when', 'where', 'why', 'how', 'each', 'any', 'both', 'some', 
        'such', 'no', 'nor', 'not', 'only', 'own', 'same', 'so', 'than', 'too', 
        'very', 'just', 'now'
    }
    
    verb_count = sum(1 for w in words if w in narrative_verbs)
    other_count = sum(1 for w in words if w in narrative_others)
    
    if verb_count >= 1 and other_count >= 2:
        standard_accs = ['property, plant and equipment', 'assets classified as held for sale',
                         'cost of materials consumed', 'changes in inventories',
                         'employee benefits expense', 'depreciation and amortisation',
                         'depreciation and amortization', 'finance costs', 'other expenses',
                         'income tax expense', 'deferred tax', 'current tax', 'earnings per share']
        for acc in standard_accs:
            if acc in text_lower:
                return False
        return True
        
    return False


def _classify_row(ws, row_idx, particulars, cy_val, py_val, is_bold):
    """
    Classifies a worksheet row into:
    - 'SIGNATURE': Auditor/Director signatures, place, date, UDIN, etc.
    - 'NOTE': Narrative sentences, comment lines, regrouping details.
    - 'FOOTER': Other metadata, page numbers, or general footers.
    - 'FINANCIAL_DATA': Genuine financial account heads.
    """
    part_lower = particulars.lower().strip()
    if not part_lower:
        return 'FOOTER'

    # 1. SIGNATURE Check
    EXACT_ROLES = {
        'partner', 'partners',
        'director', 'directors',
        'managing director',
        'cfo', 'chief financial officer',
        'company secretary',
        'auditor', 'auditors',
        'authorised signatory', 'authorized signatory'
    }

    HIGH_CONF_SUBSTR = [
        r'\bchartered\s+accountants\b',
        r'\bmembership\s+no\b',
        r'\bmembership\s+number\b',
        r'\bregistration\s+no\b',
        r'\bregistration\s+number\b',
        r'\bfirm\s+registration\s+no\b',
        r'^\s*place\s*:',
        r'^\s*date\s*:',
        r'\budin\b',
        r'\bauthorised\s+signatory\b',
        r'\bauthorized\s+signatory\b',
        r'\bfor\s+and\s+on\s+behalf\s+of\b',
        r'\bas\s+per\s+our\s+report\b',
        r'\bof\s+even\s+date\b',
        r'\bllp\b',
    ]

    EXCLUDE_WORDS = '(?:the|year|period|note|each|services|raw|other|depreciation|amortisation|amortization|interest|tax|purchase|sale|employee|allowance|provision|doubtful|bad|financial|operating|current|non-current)'
    FOR_PATTERN = re.compile(rf'^for\s+(?!{EXCLUDE_WORDS}\b)[A-Za-z0-9\s&\.,\']+(?:llp|ltd|limited|associates|co|board|directors)?$', re.I)

    for col in range(1, ws.max_column + 1):
        cell_val = ws.cell(row=row_idx, column=col).value
        if cell_val is None:
            continue
        val = str(cell_val).strip()
        if not val:
            continue
        val_lower = val.lower()

        # Check high-confidence substrings
        for p in HIGH_CONF_SUBSTR:
            if re.search(p, val_lower):
                return 'SIGNATURE'

        # Check "For <Company>"
        if FOR_PATTERN.match(val) and len(val) < 80:
            return 'SIGNATURE'

        # Check exact roles
        cleaned_val = re.sub(r'[^a-z0-9\s]', '', val_lower).strip()
        if cleaned_val in EXACT_ROLES:
            return 'SIGNATURE'

    # 2. NOTE Check
    if _is_narrative_text(particulars):
        return 'NOTE'

    return 'FINANCIAL_DATA'


def _extract_row_cells(ws, row_idx):
    row_cells = []
    for col in range(1, ws.max_column + 1):
        cell_val = ws.cell(row=row_idx, column=col).value
        row_cells.append(_clean_footer_cell(cell_val))
    return row_cells


def _clean_footer_cell(val):
    if val is None:
        return ''
    if isinstance(val, (int, float)):
        if isinstance(val, int) and 2000 <= val <= 2100:
            return str(val)
        return '-'
    val_str = str(val).strip()
    if not val_str:
        return ''
    cleaned_num = val_str.replace(',', '').replace(' ', '').replace('(', '').replace(')', '').replace('-', '').strip()
    if cleaned_num == '':
        return '-'
    try:
        float_val = float(cleaned_num)
        if float_val.is_integer() and 2000 <= int(float_val) <= 2100:
            return str(int(float_val))
        return '-'
    except ValueError:
        return val_str


def _col_letter(col_idx):
    """Convert 1-based column index to Excel letter (1='A', 2='B', etc.)."""
    result = ''
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _find_first_nonzero(data, keywords, match_type='includes'):
    """
    Search rows for the first row matching any keyword that has a non-zero CY value.
    Falls back to first matching row even if value is zero.
    Logs a warning for each zero-value match.

    Returns (cy_val, py_val).
    """
    first_match_cy = None
    first_match_py = None

    for kw in keywords:
        kw_lower = kw.lower().strip()
        for row in data:
            p = row.get('particulars', '').lower().strip()
            matched = False
            if match_type == 'exact':
                matched = (p == kw_lower)
            else:
                matched = (kw_lower in p)

            if matched:
                cy = float(row.get('cy', 0) or 0)
                py = float(row.get('py', 0) or 0)

                if cy != 0 or py != 0:
                    logger.info("Found non-zero value for '%s' on row '%s': CY=%.2f, PY=%.2f",
                                kw, row.get('particulars', ''), cy, py)
                    return cy, py

                # Record first zero match as fallback
                if first_match_cy is None:
                    first_match_cy = cy
                    first_match_py = py
                    logger.warning(
                        "Keyword '%s' matched row '%s' but value is ZERO — "
                        "may be a heading row; scanning further",
                        kw, row.get('particulars', ''))

    if first_match_cy is not None:
        logger.warning(
            "All matches for keywords %s returned zero — using zero as fallback", keywords)
        return first_match_cy, first_match_py

    logger.warning("No row found matching any of keywords: %s", keywords)
    return None, None  # Rule 5: missing ≠ zero


def _safe_float(v, default=0.0):
    """
    Safely convert a stored row value to float for arithmetic.
    Rules 5/6: None means MISSING (treated as 0.0 for ratio math only).
    Explicit 0 returns 0.0.
    """
    if v is None:
        return default
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def extract_bs_summary(bs_data):
    """
    Extract summary values from parsed BS data for ratio calculations.
    Returns dict with keys like total_equity_cy, total_ca_cy, etc.
    Uses robust multi-keyword, non-zero scanning to avoid returning 0 incorrectly.
    """
    summary = {}

    def find_value(keyword, match_type='includes'):
        """Find CY and PY values by keyword match in particulars.
        Returns (cy, py) where each may be None (missing) or float.
        Rule 5/6: None = value not present in source, not the same as 0.
        """
        kw = keyword.lower().strip()
        for row in bs_data:
            p = row.get('particulars', '').lower().strip()
            if match_type == 'exact':
                if p == kw:
                    return row.get('cy'), row.get('py')
            else:
                if kw in p:
                    return row.get('cy'), row.get('py')
        return None, None  # Rule 5: not found ≠ zero

    def find_liability_by_note(note_num, is_current):
        """Find liability values by note number and current/non-current section."""
        cy_total = 0.0
        py_total = 0.0
        in_current_section = False

        for row in bs_data:
            p = row.get('particulars', '').lower().strip()
            if 'current liabilities' in p and 'non' not in p:
                in_current_section = True
            elif 'non current liabilities' in p or 'non-current liabilities' in p:
                in_current_section = False

            note = str(row.get('note', '')).strip()
            if note == note_num and in_current_section == is_current:
                # Rule 5/6: use _safe_float so None (missing) doesn't crash addition
                cy_total += _safe_float(row.get('cy'))
                py_total += _safe_float(row.get('py'))
        
        return cy_total, py_total
    cy, py = find_value('total equity', 'exact')
    if cy is None:  # Rule 5: None = not found, try broader match
        cy, py = find_value('total equity')
    summary['total_equity_cy'] = _safe_float(cy)
    summary['total_equity_py'] = _safe_float(py)

    # Total Current Assets
    cy, py = find_value('total current assets')
    summary['total_ca_cy'] = _safe_float(cy)
    summary['total_ca_py'] = _safe_float(py)

    # Total Current Liabilities
    cy, py = find_value('total current liabilities')
    summary['total_cl_cy'] = _safe_float(cy)
    summary['total_cl_py'] = _safe_float(py)

    # Total Non-Current Liabilities
    cy, py = find_value('total non-current liabilities')
    if cy is None:  # Rule 5: None = not found, try alternate spelling
        cy, py = find_value('total non current liabilities')
    summary['total_ncl_cy'] = _safe_float(cy)
    summary['total_ncl_py'] = _safe_float(py)

    # Total Assets
    cy, py = find_value('total assets')
    summary['total_assets_cy'] = _safe_float(cy)
    summary['total_assets_py'] = _safe_float(py)

    # LTB (Long-Term Borrowings): Non-current Preference shares + Lease liabilities
    ltb_pref_cy, ltb_pref_py = find_liability_by_note('11', False)
    ltb_lease_cy, ltb_lease_py = find_liability_by_note('12', False)
    summary['ltb_cy'] = ltb_pref_cy + ltb_lease_cy
    summary['ltb_py'] = ltb_pref_py + ltb_lease_py

    # STB (Short-Term Borrowings): Current Preference shares + Lease liabilities
    stb_pref_cy, stb_pref_py = find_liability_by_note('11', True)
    stb_lease_cy, stb_lease_py = find_liability_by_note('12', True)
    summary['stb_cy'] = stb_pref_cy + stb_lease_cy
    summary['stb_py'] = stb_pref_py + stb_lease_py

    # Trade Receivables (Debtors) — scan multiple keywords, pick first non-zero row
    debtors_keywords = [
        'trade receivables',
        'sundry debtors',
        'trade and other receivables',
        'receivables',
        'debtors',
    ]
    cy, py = _find_first_nonzero(bs_data, debtors_keywords)
    summary['debtors_cy'] = _safe_float(cy)
    summary['debtors_py'] = _safe_float(py)

    if cy is None or cy == 0:  # Rule 5/6: warn if missing or genuinely zero
        logger.warning(
            "DEBTORS: Could not find non-zero trade receivables in BS — "
            "Debtor Turnover Ratio will be 0. "
            "Check that the Notes sheet contains the actual receivables figures.")

    logger.info("BS Summary: total_ca=%.0f, total_cl=%.0f, debtors=%.0f, equity=%.0f",
                summary['total_ca_cy'], summary['total_cl_cy'],
                summary['debtors_cy'], summary['total_equity_cy'])

    return summary


def extract_pl_summary(pl_data):
    """
    Extract summary values from parsed PL data for ratio calculations.
    Uses robust multi-keyword, non-zero scanning to avoid returning 0 incorrectly.
    Extracts both PBT and PAT (Profit After Tax).

    Rules 5/6/8: None = missing (not present in source), 0.0 = explicit zero.
    _safe_float() is used for arithmetic to keep ratio math stable.
    """
    summary = {}

    # ── Revenue from Operations ──
    # Try specific sub-items first, then heading rows with non-zero values
    revenue_keywords = [
        'revenue from operations',
        'net revenue from operations',
        'income from operations',
        'software services',
        'sale of products',
        'sale of services',
        'sales',
        'turnover',
        'revenue',
    ]
    cy, py = _find_first_nonzero(pl_data, revenue_keywords)
    summary['revenue_ops_cy'] = _safe_float(cy)
    summary['revenue_ops_py'] = _safe_float(py)

    if cy is None or cy == 0:  # Rule 5/6: warn for missing or genuinely zero
        logger.warning(
            "REVENUE: Could not find non-zero revenue from operations in P&L — "
            "Net Profit Ratio and Debtor Turnover Ratio will be 0. "
            "Check P&L sheet structure.")

    # ── Total Income ──
    total_income_cy, total_income_py = 0.0, 0.0
    for kw in ['total income', 'total revenue']:
        cy_t, py_t = _find_first_nonzero(pl_data, [kw])
        if cy_t is not None and cy_t != 0:  # Rule 5: None = not found
            total_income_cy, total_income_py = _safe_float(cy_t), _safe_float(py_t)
            break
    summary['total_revenue_cy'] = total_income_cy
    summary['total_revenue_py'] = total_income_py

    # ── Profit Before Tax (PBT) ──
    # Must match "profit before tax" specifically — not just any "profit" row
    pbt_cy, pbt_py = 0.0, 0.0
    pbt_keywords = [
        'profit before tax',
        'profit/(loss) before tax',
        'profit before income tax',
        'earnings before tax',
        'pbt',
    ]
    # Exact-priority search: look for rows where ALL of (profit, before, tax) appear
    for row in pl_data:
        p = row.get('particulars', '').lower().strip()
        if 'profit' in p and 'before' in p and 'tax' in p:
            cy_v = _safe_float(row.get('cy'))  # Rule 5/6: None→0.0 only for math
            py_v = _safe_float(row.get('py'))
            pbt_cy = cy_v
            pbt_py = py_v
            logger.info("PBT found: row='%s', CY=%.2f, PY=%.2f",
                        row.get('particulars', ''), cy_v, py_v)
            break
    else:
        raw_pbt_cy, raw_pbt_py = _find_first_nonzero(pl_data, pbt_keywords)
        pbt_cy = _safe_float(raw_pbt_cy)
        pbt_py = _safe_float(raw_pbt_py)
        logger.warning("PBT: exact match failed, using keyword fallback: CY=%.2f", pbt_cy)

    summary['pbt_cy'] = pbt_cy
    summary['pbt_py'] = pbt_py

    # ── Profit After Tax / Net Profit (PAT) ──
    # This is the correct numerator for Net Profit Ratio
    pat_cy, pat_py = 0.0, 0.0
    # Priority ordered: most specific first
    pat_row_candidates = [
        ('profit after tax', False),
        ('profit/(loss) after tax', False),
        ('net profit after tax', False),
        ('profit for the year', False),
        ('profit/(loss) for the year', False),
        ('net profit for the year', False),
        ('net profit', False),
        ('profit for the period', False),
        ('total comprehensive income', False),
    ]
    for kw, exact in pat_row_candidates:
        kw_lower = kw.lower()
        for row in pl_data:
            p = row.get('particulars', '').lower().strip()
            if (exact and p == kw_lower) or (not exact and kw_lower in p):
                cy_v = _safe_float(row.get('cy'))  # Rule 5/6: None→0.0 for math
                py_v = _safe_float(row.get('py'))
                if cy_v != 0 or py_v != 0:
                    pat_cy = cy_v
                    pat_py = py_v
                    logger.info("PAT found via keyword '%s': row='%s', CY=%.2f, PY=%.2f",
                                kw, row.get('particulars', ''), cy_v, py_v)
                    break
        if pat_cy != 0 or pat_py != 0:
            break

    if pat_cy == 0 and pat_py == 0:
        # Last resort: use PBT if PAT not found
        pat_cy = pbt_cy
        pat_py = pbt_py
        logger.warning(
            "PAT: Could not find Profit After Tax row — falling back to PBT (%.2f). "
            "Net Profit Ratio may be slightly overstated (ignores tax).", pat_cy)

    summary['pat_cy'] = pat_cy
    summary['pat_py'] = pat_py

    # ── Finance Charges / Finance Cost ──
    finance_keywords = ['finance charge', 'finance cost', 'interest expense',
                        'interest and finance cost', 'borrowing cost']
    cy, py = _find_first_nonzero(pl_data, finance_keywords)
    summary['finance_cost_cy'] = _safe_float(cy)
    summary['finance_cost_py'] = _safe_float(py)

    logger.info("PL Summary: revenue_ops=%.0f, pbt=%.0f, pat=%.0f, finance_cost=%.0f",
                summary['revenue_ops_cy'], summary['pbt_cy'],
                summary['pat_cy'], summary['finance_cost_cy'])

    return summary

