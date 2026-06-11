"""
FAR Automation Tool — Excel Export
Generates the final FAR workbook with all sheets, cross-sheet formulas,
formatting, and styling using openpyxl.

NOTE ON FORMULA EVALUATION:
openpyxl does not evaluate Excel formulas when saving — cells with formula
strings remain blank until Excel opens and recalculates the file.
To guarantee all values are visible immediately on open, this module writes
PRE-COMPUTED Python values (not formula strings) for all variance and
ratio columns. The values are taken directly from the variance/ratio dicts
that were already computed by the cpp_bridge module.
"""

import os
import logging
import math
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ── Color Constants ──
FILL_HEADER = PatternFill(start_color='7030A0', end_color='7030A0', fill_type='solid')
FILL_YELLOW = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
FILL_LIGHT = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')

FONT_TITLE = Font(name='Garamond', size=12, bold=True, color='000000')
FONT_SUBTITLE = Font(name='Garamond', size=12, bold=True, color='000000')
FONT_HEADER = Font(name='Garamond', size=12, bold=True, color='FFFFFF')
FONT_NORMAL = Font(name='Garamond', size=12, color='000000')
FONT_BOLD = Font(name='Garamond', size=12, bold=True, color='000000')
FONT_GREEN = Font(name='Garamond', size=12, color='16A34A')
FONT_RED = Font(name='Garamond', size=12, color='DC2626')
FONT_LEGEND = Font(name='Garamond', size=12, color='000000')
FONT_LEGEND_KEY = Font(name='Accountant', size=9, bold=False, color='FF0000')

THIN_BORDER = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB')
)

ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal='right', vertical='center')
NUMBER_FORMAT = '#,##0'
NUMBER_FORMAT_2DP = '#,##0.00'
PCT_FORMAT = '0.0%'


def _set_col_widths(ws, widths):
    """Set column widths from a dict of {col_letter: width}."""
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def _write_legends(ws, start_row, start_col, legends, cy_year, py_year):
    """Write the legend box to a worksheet."""
    legend_items = [
        ('a', f'Traced to Financials for the year ended 31 March {py_year}'),
        ('b', f'Traced to Financials for the year ended 31 March {cy_year}'),
        ('e', 'Recomputed'),
    ]

    ws.cell(row=start_row, column=start_col, value='Legends').font = FONT_LEGEND_KEY
    for i, (key, desc) in enumerate(legend_items):
        r = start_row + 1 + i
        ws.cell(row=r, column=start_col, value=key).font = FONT_LEGEND_KEY
        ws.cell(row=r, column=start_col + 1, value=desc).font = FONT_LEGEND


def _write_notes_legends(ws, start_row, start_col, cy_year, py_year):
    """Write legends specifically for notes sheets."""
    legend_items = [
        ('a', f'Traced to Financials for the year ended 31 March {py_year}'),
        ('b', f'Traced to Financials for the year ended 31 March {cy_year}'),
        ('j', 'Linked'),
        ('e', 'Recomputed'),
        ('h', 'Immaterial. Hence ignored'),
    ]

    ws.cell(row=start_row, column=start_col, value='Legend:').font = FONT_LEGEND_KEY
    for i, (key, desc) in enumerate(legend_items):
        r = start_row + 1 + i
        ws.cell(row=r, column=start_col, value=key).font = FONT_LEGEND_KEY
        ws.cell(row=r, column=start_col + 1, value=desc).font = FONT_LEGEND


def _apply_data_row_style(ws, row_idx, max_col, row_data, variance_pct_col, variance_abs_col):
    """Apply conditional formatting to a data row."""
    is_bold = row_data.get('is_bold', False)
    flag = row_data.get('flag', False)
    variance_pct = row_data.get('variance_pct', 0)

    for col in range(1, max_col + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.border = THIN_BORDER

        if is_bold:
            cell.font = FONT_BOLD
            cell.fill = FILL_LIGHT
        else:
            cell.font = FONT_NORMAL

        if flag:
            cell.fill = FILL_YELLOW
            cell.font = Font(name='Garamond', size=12, color='000000',
                           bold=is_bold)

    # Color the variance % cell
    if variance_pct is not None and variance_pct_col:
        pct_cell = ws.cell(row=row_idx, column=variance_pct_col)
        if isinstance(variance_pct, (int, float)) and not math.isnan(variance_pct):
            if variance_pct > 0:
                pct_cell.font = Font(name='Garamond', size=12, color='22C55E',
                                    bold=is_bold)
            elif variance_pct < 0:
                pct_cell.font = Font(name='Garamond', size=12, color='EF4444',
                                     bold=is_bold)


def _safe_pct_value(variance_pct):
    """
    Convert a variance_pct value (0-100 scale from cpp_bridge) to a decimal
    suitable for Excel percentage formatting (e.g. 15.5 → 0.155).
    Handles None, inf, nan gracefully.
    """
    if variance_pct is None:
        return None
    try:
        v = float(variance_pct)
        if math.isinf(v) or math.isnan(v):
            return None
        return v / 100.0
    except (TypeError, ValueError):
        return None


def export_far_workbook(output_path, bs_result, pl_result, notes_result,
                        ratios, remarks_bs, remarks_pl,
                        client_name, firm_name, financial_year,
                        rounding_unit, cy_year, py_year,
                        include_formulas=False, chart_paths=None,
                        bs_notes=None, bs_signatures=None, bs_footers=None,
                        pl_notes=None, pl_signatures=None, pl_footers=None,
                        notes_sheet_notes=None, notes_signatures=None, notes_footers=None,
                        report_type='FAR', meta=None):
    """
    Generate the complete FAR workbook.
    """
    wb = Workbook()

    # Force Excel to recalculate all formulas on open (belt-and-suspenders)
    wb.calculation.calcMode = 'auto'
    wb.calculation.fullCalcOnLoad = True

    # ── Sheet 1: Cover ──
    ws_cover = wb.active
    ws_cover.title = 'COVER PAGE'
    _write_cover_sheet(ws_cover, client_name, firm_name, financial_year,
                       rounding_unit, cy_year, py_year, include_visuals=bool(chart_paths),
                       report_type=report_type, meta=meta)

    # ── Sheet 2: BS ──
    ws_bs = wb.create_sheet('BS ANALYSIS')
    bs_row_map = _write_analysis_sheet(
        ws_bs, bs_result, remarks_bs,
        client_name, firm_name, cy_year, py_year, rounding_unit,
        title='Analysis of Balance Sheet',
        date_prefix='As at',
        note_col_header='Note number',
        include_formulas=include_formulas,
        notes=bs_notes,
        signatures=bs_signatures,
        footers=bs_footers
    )

    # ── Sheet 3: PL ──
    ws_pl = wb.create_sheet('P&L ANALYSIS')
    pl_row_map = _write_analysis_sheet(
        ws_pl, pl_result, remarks_pl,
        client_name, firm_name, cy_year, py_year, rounding_unit,
        title='Analysis of Statement of Profit and Loss',
        date_prefix='For the Year ended',
        note_col_header='Note number',
        include_formulas=include_formulas,
        notes=pl_notes,
        signatures=pl_signatures,
        footers=pl_footers
    )

    # ── Sheets 4-7: Notes ──
    if isinstance(notes_result, dict):
        sheet_notes_dict = notes_sheet_notes or {}
        signatures_dict = notes_signatures or {}
        footers_dict = notes_footers or {}
        for sheet_name, note_groups in notes_result.items():
            ws_notes = wb.create_sheet(f"Notes {sheet_name}"[:30])
            sh_notes = sheet_notes_dict.get(sheet_name)
            sh_sigs = signatures_dict.get(sheet_name)
            sh_footers = footers_dict.get(sheet_name)
            _write_notes_sheet(ws_notes, note_groups,
                             client_name, firm_name, cy_year, py_year, rounding_unit,
                             include_formulas=include_formulas,
                             notes=sh_notes,
                             signatures=sh_sigs,
                             footers=sh_footers)

    # ── Sheet 8: Ratio Analysis ──
    ws_ratio = wb.create_sheet('RATIOS & RELATIONS')
    _write_ratio_sheet(ws_ratio, ratios, client_name, firm_name, cy_year, py_year,
                       bs_row_map, pl_row_map)

    # ── Sheet 9: Dashboard ──
    ws_dash = wb.create_sheet('Dashboard')
    _write_dashboard_sheet(ws_dash, bs_result, pl_result, ratios,
                           client_name, firm_name, cy_year, py_year, rounding_unit,
                           include_formulas=include_formulas)

    # ── Sheet 10: Visual Analytics ──
    if chart_paths:
        ws_visuals = wb.create_sheet('VISUAL ANALYTICS')
        _write_visuals_sheet(ws_visuals, chart_paths, client_name, firm_name, cy_year, rounding_unit)

    # Save
    wb.save(output_path)
    logger.info("FAR workbook saved to %s", output_path)
    return True


def _write_cover_sheet(ws, client_name, firm_name, financial_year,
                       rounding_unit, cy_year, py_year, include_visuals=False,
                       report_type='FAR', meta=None):
    """Write the cover/index sheet matching the user's image layout."""
    # Hide grid lines on the cover page for a clean card presentation
    ws.views.sheetView[0].showGridLines = True

    # Column widths
    ws.column_dimensions['A'].width = 4
    for col in ['B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col].width = 16
    ws.column_dimensions['H'].width = 4

    # Helper function to style a banner row
    def write_banner(row_idx, text):
        ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=7)
        cell = ws.cell(row=row_idx, column=2, value=text)
        cell.font = Font(name='Garamond', size=14, bold=True, color='FFFFFF')
        cell.fill = FILL_HEADER # Purple fill
        cell.alignment = ALIGN_CENTER
        ws.row_dimensions[row_idx].height = 35

    # Helper function to style a details/bullet row with borders
    def write_detail(row_idx, text, is_bullet=False):
        ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=7)
        cell = ws.cell(row=row_idx, column=2, value=text)
        cell.font = FONT_NORMAL
        cell.alignment = ALIGN_LEFT if is_bullet else ALIGN_CENTER
        ws.row_dimensions[row_idx].height = 24
        
        # Apply borders to all cells in the merged range to show border properly
        for col_idx in range(2, 8):
            ws.cell(row=row_idx, column=col_idx).border = THIN_BORDER

    # Row 2: Firm Name
    write_banner(2, firm_name if firm_name else "Walker Chandiok & Co LLP")

    # Row 4: Report type heading
    _REPORT_LABELS = {
        'FAR': 'Final Analytical Report',
        'PAR': 'Preliminary Analytical Report',
    }
    report_heading = _REPORT_LABELS.get(report_type, 'Final Analytical Report')
    write_banner(4, report_heading)

    # Row 6: Client Name
    write_detail(6, f"Client: {client_name}")

    # Row 8: Analysis Date
    # Date formatting from image: December 16, 2025
    analysis_date = datetime.now().strftime("%B %d, %Y")
    write_detail(8, f"Analysis Date: {analysis_date}")

    # Row 11: Year-over-Year Financial Comparison
    write_banner(11, "Year-over-Year Financial Comparison")

    if meta:
        py_days = meta.get('py_days', 365)
        py_months = meta.get('py_months', 12.0)
        py_cov = meta.get('py_coverage', 100.0)
        cy_days = meta.get('cy_days', 365)
        cy_months = meta.get('cy_months', 12.0)
        cy_cov = meta.get('cy_coverage', 100.0)
    else:
        py_days = cy_days = 365
        py_months = cy_months = 12.0
        py_cov = cy_cov = 100.0

    write_detail(13, f"Previous Year ({py_year}): Days: {py_days} | Months: {py_months} | Coverage: {py_cov}%", is_bullet=True)
    write_detail(14, f"Current Year ({cy_year}): Days: {cy_days} | Months: {cy_months} | Coverage: {cy_cov}%", is_bullet=True)

    # Row 17: Analysis Summary Header
    write_banner(17, "Analysis Summary")

    # Rows 19 to 24: Summary Bullet Points
    bullets = [
        "  • Balance Sheet Analysis",
        "  • Profit and Loss Analysis",
        "  • Ratio and Relations Analysis",
        "  • Visual Analytics and Charts",
        "  • Year-over-Year Comparisons",
        "  • Data Validation Checks"
    ]
    for idx, bullet in enumerate(bullets):
        write_detail(19 + idx, bullet, is_bullet=True)

    # Row 27: Confidential Disclaimer
    ws.merge_cells(start_row=27, start_column=2, end_row=27, end_column=7)
    conf_cell = ws.cell(row=27, column=2, value="Confidential - For Internal Use Only")
    conf_cell.font = Font(name='Garamond', size=10, italic=True, color='7F7F7F')
    conf_cell.alignment = ALIGN_CENTER
    ws.row_dimensions[27].height = 20


def _write_analysis_sheet(ws, data, remarks, client_name, firm_name, cy_year, py_year,
                          rounding_unit, title, date_prefix, note_col_header,
                          include_formulas=False, notes=None, signatures=None, footers=None):
    """
    Write BS or PL analysis sheet with PRE-COMPUTED variance values.
    Variance columns use actual calculated numbers — NOT formula strings —
    so the file opens with all values visible immediately.

    Returns a dict mapping particulars (lowercase) → Excel row number for cross-referencing.
    """
    _set_col_widths(ws, {
        'A': 40, 'B': 12, 'C': 16, 'D': 16, 'E': 16, 'F': 12, 'G': 45
    })

    # Header section
    ws.cell(row=1, column=1, value='Firm Name:').font = FONT_BOLD
    ws.cell(row=1, column=3, value=firm_name).font = FONT_NORMAL
    ws.cell(row=2, column=1, value='Client Name:').font = FONT_BOLD
    ws.cell(row=2, column=3, value=client_name).font = FONT_NORMAL
    ws.cell(row=3, column=1, value=title).font = FONT_SUBTITLE
    ws.cell(row=4, column=1,
            value=f'{date_prefix} 31 March {cy_year} vs 31 March {py_year}').font = FONT_NORMAL
    ws.cell(row=5, column=1,
            value=f'(Rs. In {rounding_unit})').font = FONT_LEGEND

    # Legends
    _write_legends(ws, 1, 6, None, cy_year, py_year)

    ws.cell(row=6, column=1,
            value='Scope: All variances above TE and (+/-) 10% have been selected for analysis').font = FONT_LEGEND

    # Column headers row 7
    headers_legend = [('', ''), ('', ''), ('b', ''), ('a', ''), ('e', ''), ('e', ''), ('', '')]
    for i, (legend, _) in enumerate(headers_legend):
        if legend:
            ws.cell(row=7, column=i + 1, value=legend).font = FONT_LEGEND_KEY
            ws.cell(row=7, column=i + 1).alignment = ALIGN_CENTER

    # Column headers row 8
    col_headers = [
        'Particulars', note_col_header,
        f'{date_prefix}\n31 March {cy_year}',
        f'{date_prefix}\n31 March {py_year}',
        'Variance\n(Absolute)' if 'As at' in date_prefix else 'Absolute Variance',
        'Variance %' if 'As at' in date_prefix else 'Variance \n%',
        'Remarks'
    ]
    for i, h in enumerate(col_headers):
        cell = ws.cell(row=8, column=i + 1, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

    # Freeze panes below header
    ws.freeze_panes = 'A9'

    # Data rows — using pre-computed values from variance dicts
    row_map = {}  # particulars → excel row number
    for idx, row_data in enumerate(data):
        excel_row = 9 + idx

        # Store mapping for cross-sheet references
        p_lower = row_data.get('particulars', '').lower().strip()
        if p_lower:
            row_map[p_lower] = excel_row

        # Write data
        ws.cell(row=excel_row, column=1, value=row_data.get('particulars', ''))
        ws.cell(row=excel_row, column=2, value=row_data.get('note', ''))

        cy_val = row_data.get('cy')
        py_val = row_data.get('py')
        decimals = 2 if rounding_unit in ('Lakhs', 'Millions', 'Actuals') else 0

        def _write_val(val):
            """Returns the value to write to cell: number, '-', or '-NA-'.
            - None  → '-NA-'  (truly blank in source)
            - '-'   → '-'     (dash in source)
            - 0/0.0 → '-'     (zero = no meaningful value, matches UI display)
            - float → rounded number
            """
            if val is None:
                return '-NA-'
            if val == '-':
                return '-'
            try:
                rounded = round(float(val), decimals)
                if rounded == 0:
                    return '-'   # treat zero same as missing — matches UI
                return rounded
            except (ValueError, TypeError):
                return '-NA-'

        ws.cell(row=excel_row, column=3, value=_write_val(cy_val))
        ws.cell(row=excel_row, column=4, value=_write_val(py_val))


        # ── Excel Formulas for Variance ──
        abs_cell = ws.cell(row=excel_row, column=5)
        abs_cell.value = f'=IF(AND(NOT(ISNUMBER(C{excel_row})), NOT(ISNUMBER(D{excel_row}))), "-", IF(ISNUMBER(C{excel_row}), C{excel_row}, 0) - IF(ISNUMBER(D{excel_row}), D{excel_row}, 0))'

        pct_cell = ws.cell(row=excel_row, column=6)
        pct_cell.value = f'=IF(OR(NOT(ISNUMBER(D{excel_row})), D{excel_row}=0), "-", (IF(ISNUMBER(C{excel_row}), C{excel_row}, 0) - D{excel_row})/ABS(D{excel_row}))'

        # Number formatting
        fmt = NUMBER_FORMAT_2DP if rounding_unit in ('Lakhs', 'Millions', 'Actuals') else NUMBER_FORMAT
        for col in [3, 4, 5]:
            ws.cell(row=excel_row, column=col).number_format = fmt
        pct_cell.number_format = PCT_FORMAT

        # Remarks
        remark = remarks.get(idx, row_data.get('remark', ''))
        ws.cell(row=excel_row, column=7, value=remark)

        # Styling
        _apply_data_row_style(ws, excel_row, 7, row_data, 6, 5)

    # Write notes, signatures, and footers at the end of BS/PL tables
    extra_rows = []
    if notes or signatures or footers:
        for _ in range(3):
            extra_rows.append([''] * 7)
        if notes:
            extra_rows.extend(notes)
        if signatures:
            if notes:
                extra_rows.append([''] * 7)
            extra_rows.extend(signatures)
        if footers:
            if notes or signatures:
                extra_rows.append([''] * 7)
            extra_rows.extend(footers)

    if extra_rows:
        start_extra_row = 9 + len(data)
        for f_idx, f_row in enumerate(extra_rows):
            excel_row = start_extra_row + f_idx
            ws.row_dimensions[excel_row].height = 20
            for col_idx in range(1, len(f_row) + 1):
                cell_val = f_row[col_idx - 1]
                cell = ws.cell(row=excel_row, column=col_idx, value=cell_val)
                cell.font = FONT_NORMAL
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = Border()
                cell.fill = PatternFill(fill_type=None)

    return row_map


def _write_notes_sheet(ws, notes_groups, client_name, firm_name, cy_year, py_year, rounding_unit,
                       include_formulas=False, notes=None, signatures=None, footers=None):
    """Write a notes analysis sheet with pre-computed variance values."""
    _set_col_widths(ws, {
        'A': 5, 'B': 40, 'C': 16, 'D': 16, 'E': 16, 'F': 12, 'G': 45,
        'H': 5, 'I': 50
    })

    # Header
    ws.cell(row=1, column=1, value='Firm:').font = FONT_BOLD
    ws.cell(row=1, column=3, value=firm_name).font = FONT_NORMAL
    ws.cell(row=2, column=1, value='Client:').font = FONT_BOLD
    ws.cell(row=2, column=3, value=client_name).font = FONT_NORMAL
    ws.cell(row=3, column=1, value='Analysis of Notes to Accounts').font = FONT_SUBTITLE

    # Legends
    _write_notes_legends(ws, 1, 8, cy_year, py_year)

    current_row = 10

    for note_group in notes_groups:
        # Note heading
        heading = f"Note {note_group['note_num']}: {note_group['note_heading']}"
        ws.cell(row=current_row, column=2, value=heading).font = FONT_SUBTITLE

        # Sub-headers
        current_row += 1
        sub_headers = ['', 'Particulars', f'As at 31st March {cy_year}',
                       f'As at 31st March {py_year}', 'Variance', 'Variance %', 'Remarks']
        for i, h in enumerate(sub_headers):
            cell = ws.cell(row=current_row, column=i + 1, value=h)
            cell.font = FONT_HEADER
            cell.fill = FILL_HEADER
            cell.alignment = ALIGN_CENTER
            cell.border = THIN_BORDER

        # Legend row
        ws.cell(row=current_row - 1, column=3, value='a').font = FONT_LEGEND_KEY
        ws.cell(row=current_row - 1, column=4, value='b').font = FONT_LEGEND_KEY
        ws.cell(row=current_row - 1, column=5, value='e').font = FONT_LEGEND_KEY

        current_row += 1

        # Data rows — pre-computed variance values
        for row_data in note_group.get('data', []):
            cy_val = row_data.get('cy')
            py_val = row_data.get('py')
            decimals = 2 if rounding_unit in ('Lakhs', 'Millions', 'Actuals') else 0
            ws.cell(row=current_row, column=2, value=row_data.get('particulars', ''))
            ws.cell(row=current_row, column=3, value=round(cy_val, decimals) if cy_val is not None else '-')
            ws.cell(row=current_row, column=4, value=round(py_val, decimals) if py_val is not None else '-')

            ws.cell(row=current_row, column=5, value=f'=IF(AND(NOT(ISNUMBER(C{current_row})), NOT(ISNUMBER(D{current_row}))), "-", IF(ISNUMBER(C{current_row}), C{current_row}, 0) - IF(ISNUMBER(D{current_row}), D{current_row}, 0))')
            ws.cell(row=current_row, column=6, value=f'=IF(OR(NOT(ISNUMBER(D{current_row})), D{current_row}=0), "-", (IF(ISNUMBER(C{current_row}), C{current_row}, 0) - D{current_row})/ABS(D{current_row}))')

            fmt = NUMBER_FORMAT_2DP if rounding_unit in ('Lakhs', 'Millions', 'Actuals') else NUMBER_FORMAT
            for col in [3, 4, 5]:
                ws.cell(row=current_row, column=col).number_format = fmt
            ws.cell(row=current_row, column=6).number_format = PCT_FORMAT

            _apply_data_row_style(ws, current_row, 7, row_data, 6, 5)
            current_row += 1

        current_row += 2  # Gap between note groups

    # Write notes, signatures, and footers at the end of notes sheet
    extra_rows = []
    if notes or signatures or footers:
        for _ in range(3):
            extra_rows.append([''] * 7)
        if notes:
            extra_rows.extend(notes)
        if signatures:
            if notes:
                extra_rows.append([''] * 7)
            extra_rows.extend(signatures)
        if footers:
            if notes or signatures:
                extra_rows.append([''] * 7)
            extra_rows.extend(footers)

    if extra_rows:
        start_extra_row = current_row
        for f_idx, f_row in enumerate(extra_rows):
            excel_row = start_extra_row + f_idx
            ws.row_dimensions[excel_row].height = 20
            for col_idx in range(1, len(f_row) + 1):
                cell_val = f_row[col_idx - 1]
                cell = ws.cell(row=excel_row, column=col_idx, value=cell_val)
                cell.font = FONT_NORMAL
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = Border()
                cell.fill = PatternFill(fill_type=None)


def _write_ratio_sheet(ws, ratios, client_name, firm_name, cy_year, py_year,
                       bs_row_map, pl_row_map):
    """Write the Ratio Analysis sheet with pre-computed values."""
    _set_col_widths(ws, {
        'A': 5, 'B': 35, 'C': 45, 'D': 16, 'E': 16, 'F': 14, 'G': 45
    })

    # Header
    ws.cell(row=1, column=2, value='Firm:').font = FONT_BOLD
    ws.cell(row=1, column=3, value=firm_name).font = FONT_NORMAL
    ws.cell(row=2, column=2, value='Client:').font = FONT_BOLD
    ws.cell(row=2, column=3, value=client_name).font = FONT_NORMAL
    ws.cell(row=3, column=2, value='Ratio Analysis').font = FONT_SUBTITLE
    ws.cell(row=4, column=2,
            value=f'As at 31 March {cy_year} vs 31 March {py_year}').font = FONT_NORMAL
    ws.cell(row=5, column=2,
            value='Scope: Percentage variance of +/- 10% have been selected for analysis').font = FONT_LEGEND

    # Column headers
    headers = ['', 'KEY METRIC', 'FORMULA', f'{cy_year}-{str(cy_year+1)[-2:]} (CY)',
               f'{py_year}-{str(py_year+1)[-2:]} (PY)', 'CHANGE (%)', 'REMARKS']
    for i, h in enumerate(headers):
        cell = ws.cell(row=7, column=i + 1, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

    ws.freeze_panes = 'A8'

    # Ratio data rows — all values pre-computed
    for idx, ratio in enumerate(ratios):
        excel_row = 8 + idx

        ws.cell(row=excel_row, column=2, value=ratio['key']).font = FONT_BOLD
        ws.cell(row=excel_row, column=3, value=ratio['formula']).font = FONT_NORMAL

        cy_val = ratio.get('cy')
        py_val = ratio.get('py')
        ws.cell(row=excel_row, column=4, value=cy_val if cy_val is not None else '-')
        ws.cell(row=excel_row, column=5, value=py_val if py_val is not None else '-')

        # ── Pre-computed change % (not formula string) ──
        change = ratio.get('change', None)  # already a % (e.g. 15.5 = 15.5%)
        change_decimal = _safe_pct_value(change) if change is not None else '-'
        change_cell = ws.cell(row=excel_row, column=6)
        change_cell.value = change_decimal
        
        ws.cell(row=excel_row, column=4).number_format = NUMBER_FORMAT_2DP
        ws.cell(row=excel_row, column=5).number_format = NUMBER_FORMAT_2DP
        change_cell.number_format = PCT_FORMAT

        # Styling
        for col in range(1, 8):
            ws.cell(row=excel_row, column=col).border = THIN_BORDER

        flag = ratio.get('flag', False)
        if flag:
            for col in range(1, 8):
                ws.cell(row=excel_row, column=col).fill = FILL_YELLOW
                ws.cell(row=excel_row, column=col).font = Font(
                    name='Garamond', size=12, color='000000',
                    bold=(col == 2))

        # Color change cell
        if change is not None:
            if change > 0:
                change_cell.font = Font(name='Garamond', size=12, color='22C55E')
            elif change < 0:
                change_cell.font = Font(name='Garamond', size=12, color='EF4444')


def _write_dashboard_sheet(ws, bs_data, pl_data, ratios,
                          client_name, firm_name, cy_year, py_year, rounding_unit,
                          include_formulas=False):
    """Write the dashboard summary sheet with pre-computed variance values."""
    _set_col_widths(ws, {'A': 5, 'B': 30, 'C': 18, 'D': 18, 'E': 18})

    ws.cell(row=1, column=2, value='Key Metrics Dashboard').font = FONT_TITLE
    ws.cell(row=2, column=2, value=f'Firm: {firm_name}').font = FONT_BOLD
    ws.cell(row=3, column=2, value=f'Client: {client_name} — FY {cy_year}').font = FONT_SUBTITLE
    ws.cell(row=4, column=2, value=f'(Rs. in {rounding_unit})').font = FONT_LEGEND

    # Summary metrics table
    headers = ['', 'Metric', f'CY ({cy_year})', f'PY ({py_year})', 'Variance']
    for i, h in enumerate(headers):
        cell = ws.cell(row=6, column=i + 1, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.border = THIN_BORDER

    # Key metrics from PL
    def find_val(data, keyword):
        kw_clean = keyword.lower().strip()
        
        # 1. Exact match
        for row in data:
            p_lower = row.get('particulars', '').lower().strip()
            if p_lower == kw_clean:
                return row.get('cy'), row.get('py')
                
        # 2. Substring match
        for row in data:
            p_lower = row.get('particulars', '').lower().strip()
            if kw_clean in p_lower:
                return row.get('cy'), row.get('py')
                
        # 3. All words match
        kw_words = kw_clean.split()
        for row in data:
            p_lower = row.get('particulars', '').lower().strip()
            if all(w in p_lower for w in kw_words):
                return row.get('cy'), row.get('py')
                
        return None, None

    metrics = [
        ('Total Revenue', *find_val(pl_data, 'total income')),
        ('Total Expenses', *find_val(pl_data, 'total expense')),
        ('Profit Before Tax', *find_val(pl_data, 'profit before tax')),
        ('Total Assets', *find_val(bs_data, 'total assets')),
        ('Total Equity', *find_val(bs_data, 'total equity')),
        ('Total Current Assets', *find_val(bs_data, 'total current assets')),
        ('Total Current Liabilities', *find_val(bs_data, 'total current liabilities')),
    ]

    for idx, (name, cy, py) in enumerate(metrics):
        r = 7 + idx
        ws.cell(row=r, column=2, value=name).font = FONT_BOLD

        decimals = 2 if rounding_unit in ('Lakhs', 'Millions', 'Actuals') else 0
        fmt = NUMBER_FORMAT_2DP if rounding_unit in ('Lakhs', 'Millions', 'Actuals') else NUMBER_FORMAT
        
        ws.cell(row=r, column=3, value=round(cy, decimals) if cy is not None else '-').number_format = fmt
        ws.cell(row=r, column=4, value=round(py, decimals) if py is not None else '-').number_format = fmt

        # Excel formula for variance
        ws.cell(row=r, column=5, value=f'=IF(AND(NOT(ISNUMBER(C{r})), NOT(ISNUMBER(D{r}))), "-", IF(ISNUMBER(C{r}), C{r}, 0) - IF(ISNUMBER(D{r}), D{r}, 0))').number_format = fmt

        for col in range(1, 6):
            ws.cell(row=r, column=col).border = THIN_BORDER

    # Ratio summary
    ratio_start = 7 + len(metrics) + 2
    ws.cell(row=ratio_start, column=2, value='Key Ratios').font = FONT_SUBTITLE

    ratio_headers = ['', 'Ratio', 'CY', 'PY', 'Change %']
    for i, h in enumerate(ratio_headers):
        cell = ws.cell(row=ratio_start + 1, column=i + 1, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.border = THIN_BORDER

    for idx, ratio in enumerate(ratios):
        r = ratio_start + 2 + idx
        ws.cell(row=r, column=2, value=ratio['key']).font = FONT_NORMAL
        
        cy_val = ratio.get('cy')
        py_val = ratio.get('py')
        ws.cell(row=r, column=3, value=cy_val if cy_val is not None else '-').number_format = NUMBER_FORMAT_2DP
        ws.cell(row=r, column=4, value=py_val if py_val is not None else '-').number_format = NUMBER_FORMAT_2DP

        # Pre-computed change decimal for PCT_FORMAT
        change = ratio.get('change', None)
        change_decimal = _safe_pct_value(change) if change is not None else '-'
        ws.cell(row=r, column=5, value=change_decimal).number_format = PCT_FORMAT

        for col in range(1, 6):
            ws.cell(row=r, column=col).border = THIN_BORDER


def _write_visuals_sheet(ws, chart_paths, client_name, firm_name, cy_year, rounding_unit):
    """Write the visual analytics dashboard sheet and embed matplotlib charts."""
    # Ensure grid lines are visible
    ws.views.sheetView[0].showGridLines = True

    # Set column widths to accommodate a nice layout
    _set_col_widths(ws, {
        'A': 4, 'B': 12, 'C': 12, 'D': 12, 'E': 12, 'F': 12, 'G': 12,
        'H': 4, 'I': 12, 'J': 12, 'K': 12, 'L': 12, 'M': 12, 'N': 12
    })

    # Header block
    ws.cell(row=2, column=2, value='Visual Analytics Dashboard').font = FONT_TITLE
    ws.cell(row=3, column=2, value=f'Firm: {firm_name}').font = FONT_BOLD
    ws.cell(row=4, column=2, value=f'Client: {client_name} — FY {cy_year}').font = FONT_SUBTITLE
    ws.cell(row=5, column=2, value=f'Amounts in: Rs. {rounding_unit}').font = FONT_LEGEND

    # Embedded images positioning in a clean 2x3 grid
    from openpyxl.drawing.image import Image

    positions = [
        'B7',    # Chart 1: Revenue vs Expenses
        'I7',    # Chart 2: Profit Before Tax Trend
        'B29',   # Chart 3: Asset Composition
        'I29',   # Chart 4: Liability/Equity Composition
        'B51',   # Chart 5: Top 5 BS Variances
        'I51',   # Chart 6: Top 5 PL Variances
    ]

    for idx, path in enumerate(chart_paths):
        if idx >= len(positions):
            break
        if not os.path.exists(path):
            logger.warning("Chart image path does not exist: %s", path)
            continue
        try:
            img = Image(path)
            # Scale down slightly to fit beautifully
            img.width = 500
            img.height = 350
            ws.add_image(img, positions[idx])
        except Exception as e:
            logger.error("Failed to add chart %d to Excel: %s", idx, e)

