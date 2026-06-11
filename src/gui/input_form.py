"""
FAR Automation Tool — Input Form (Module 1)
Client information form with file upload panels and Generate button.
"""

import os
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QFormLayout, QGroupBox,
    QLineEdit, QComboBox, QRadioButton, QCheckBox, QPushButton,
    QLabel, QFileDialog, QProgressDialog, QButtonGroup, QFrame,
    QScrollArea, QSizePolicy, QMessageBox, QDateEdit
)
from PyQt6.QtCore import pyqtSignal, Qt, QThread, pyqtSlot, QDate
from PyQt6.QtGui import QFont


class GenerateWorker(QThread):
    """Background thread for FAR generation."""
    progress = pyqtSignal(int, str)
    finished = pyqtSignal(bool, str)
    
    def __init__(self, callback):
        super().__init__()
        self.callback = callback
    
    def run(self):
        try:
            self.callback(self.progress.emit)
            self.finished.emit(True, "FAR generated successfully!")
        except Exception as e:
            self.finished.emit(False, str(e))


class FileUploadPanel(QGroupBox):
    """A file upload panel with Browse button and file info display."""
    
    fileSelected = pyqtSignal(str)
    
    def __init__(self, title, multi_select=False, parent=None):
        super().__init__(title, parent)
        self.multi_select = multi_select
        self.file_paths = []
        self._setup_ui()
    
    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(8)
        
        # Browse button row
        btn_row = QHBoxLayout()
        self.browse_btn = QPushButton('📂 Browse...')
        self.browse_btn.setFixedWidth(140)
        self.browse_btn.clicked.connect(self._browse)
        btn_row.addWidget(self.browse_btn)
        
        self.status_icon = QLabel('⬜')
        self.status_icon.setObjectName('statusIcon')
        btn_row.addWidget(self.status_icon)
        
        btn_row.addStretch()
        layout.addLayout(btn_row)
        
        # File path display
        self.path_label = QLabel('No file selected')
        self.path_label.setObjectName('mutedLabel')
        self.path_label.setWordWrap(True)
        layout.addWidget(self.path_label)
        
    
    def _browse(self):
        if self.multi_select:
            files, _ = QFileDialog.getOpenFileNames(
                self, 'Select Excel Files', '',
                'Excel Files (*.xlsx *.xls)')
            if files:
                self.file_paths = files
                self._update_display()
        else:
            file_path, _ = QFileDialog.getOpenFileName(
                self, 'Select Excel File', '',
                'Excel Files (*.xlsx *.xls)')
            if file_path:
                self.file_paths = [file_path]
                self._update_display()
    
    def _update_display(self):
        if self.file_paths:
            names = [os.path.basename(f) for f in self.file_paths]
            self.path_label.setText(', '.join(names))
            self.path_label.setObjectName('successLabel')
            self.status_icon.setText('✅')
            

            self.fileSelected.emit(self.file_paths[0])
    
    def get_paths(self):
        return self.file_paths
    
    def is_loaded(self):
        return len(self.file_paths) > 0


class InputForm(QWidget):
    """Module 1 — Input Form for client information and file uploads."""
    
    generateRequested = pyqtSignal(dict)
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self._setup_ui()
    
    def _setup_ui(self):
        # Scroll area wrapper
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        
        content = QWidget()
        main_layout = QVBoxLayout(content)
        main_layout.setContentsMargins(32, 24, 32, 24)
        main_layout.setSpacing(20)
        
        # Title
        title = QLabel('📝 Input Form')
        title.setObjectName('headingLabel')
        main_layout.addWidget(title)
        
        subtitle = QLabel('Enter client details and upload financial statements')
        subtitle.setObjectName('mutedLabel')
        main_layout.addWidget(subtitle)
        
        # ── Client Information ──
        info_group = QGroupBox('Client Information')
        info_layout = QFormLayout(info_group)
        info_layout.setSpacing(12)
        info_layout.setContentsMargins(16, 20, 16, 16)
        
        self.firm_name = QLineEdit()
        self.firm_name.setPlaceholderText('Enter firm name...')
        info_layout.addRow('Firm Name:', self.firm_name)
        
        self.client_name = QLineEdit()
        self.client_name.setPlaceholderText('Enter client name...')
        info_layout.addRow('Client Name:', self.client_name)
        
        # YoY Comparison configuration
        comp_group = QGroupBox('Year-over-Year Comparison Settings')
        comp_main_layout = QVBoxLayout(comp_group)
        comp_main_layout.setSpacing(12)
        
        cols_layout = QHBoxLayout()
        
        # Previous Year Column
        py_widget = QWidget()
        py_layout = QFormLayout(py_widget)
        py_layout.setSpacing(8)
        py_layout.setContentsMargins(0, 0, 0, 0)
        
        self.py_year_sel = QComboBox()
        years = [f'{y}-{str(y+1)[-2:]}' for y in range(2025, 2019, -1)]
        self.py_year_sel.addItems(years)
        if '2024-25' in years:
            self.py_year_sel.setCurrentText('2024-25')
        else:
            self.py_year_sel.setCurrentIndex(1)
            
        self.py_from_date = QLineEdit()
        self.py_from_date.setPlaceholderText('dd-mm-yyyy')
        self.py_from_date.setText('01-04-2024')
        self.py_from_date.textChanged.connect(self._update_metrics)
        
        self.py_to_date = QLineEdit()
        self.py_to_date.setPlaceholderText('dd-mm-yyyy')
        self.py_to_date.setText('31-03-2025')
        self.py_to_date.textChanged.connect(self._update_metrics)
        
        py_layout.addRow('PY Year Selector:', self.py_year_sel)
        py_layout.addRow('PY From Date:', self.py_from_date)
        py_layout.addRow('PY To Date:', self.py_to_date)
        
        self.py_metrics_label = QLabel('')
        self.py_metrics_label.setObjectName('mutedLabel')
        py_layout.addRow('', self.py_metrics_label)
        
        cols_layout.addWidget(py_widget)
        
        # Vertical Separator
        v_sep = QFrame()
        v_sep.setFrameShape(QFrame.Shape.VLine)
        v_sep.setStyleSheet("color: #E0E0E0;")
        cols_layout.addWidget(v_sep)
        
        # Current Year Column
        cy_widget = QWidget()
        cy_layout = QFormLayout(cy_widget)
        cy_layout.setSpacing(8)
        cy_layout.setContentsMargins(0, 0, 0, 0)
        
        self.cy_year_sel = QComboBox()
        self.cy_year_sel.addItems(years)
        self.cy_year_sel.setCurrentText('2025-26')
        
        self.cy_from_date = QLineEdit()
        self.cy_from_date.setPlaceholderText('dd-mm-yyyy')
        self.cy_from_date.setText('01-04-2025')
        self.cy_from_date.textChanged.connect(self._update_metrics)
        
        self.cy_to_date = QLineEdit()
        self.cy_to_date.setPlaceholderText('dd-mm-yyyy')
        self.cy_to_date.setText('31-03-2026')
        self.cy_to_date.textChanged.connect(self._update_metrics)
        
        cy_layout.addRow('CY Year Selector:', self.cy_year_sel)
        cy_layout.addRow('CY From Date:', self.cy_from_date)
        cy_layout.addRow('CY To Date:', self.cy_to_date)
        
        self.cy_metrics_label = QLabel('')
        self.cy_metrics_label.setObjectName('mutedLabel')
        cy_layout.addRow('', self.cy_metrics_label)
        
        cols_layout.addWidget(cy_widget)
        comp_main_layout.addLayout(cols_layout)
        
        h_sep = QFrame()
        h_sep.setFrameShape(QFrame.Shape.HLine)
        h_sep.setStyleSheet("color: #E0E0E0;")
        comp_main_layout.addWidget(h_sep)
        
        toggle_layout = QHBoxLayout()
        self.auto_annualize_cb = QCheckBox("Auto Annualize Partial-Year Data")
        self.auto_annualize_cb.setChecked(True)
        self.auto_annualize_cb.toggled.connect(self._update_metrics)
        toggle_layout.addWidget(self.auto_annualize_cb)
        
        self.warning_label = QLabel("⚠️ Selected periods are different. Results are being normalized to a full-year equivalent basis.")
        self.warning_label.setStyleSheet("color: #EAB308; font-weight: 600; font-size: 11px;")
        self.warning_label.setVisible(False)
        toggle_layout.addWidget(self.warning_label)
        toggle_layout.addStretch()
        
        comp_main_layout.addLayout(toggle_layout)
        
        info_layout.addRow('', comp_group)
        self._update_metrics()
        
        # ── Report Type ──
        report_type_widget = QWidget()
        rt_layout = QHBoxLayout(report_type_widget)
        rt_layout.setContentsMargins(0, 0, 0, 0)
        rt_layout.setSpacing(24)
        self.report_type_group = QButtonGroup()

        self.rb_far = QRadioButton("FAR  (Final Analytical Report)")
        self.rb_far.setChecked(True)
        self.rb_far.setObjectName('reportTypeRadio')
        self.report_type_group.addButton(self.rb_far, 0)
        rt_layout.addWidget(self.rb_far)

        self.rb_par = QRadioButton("PAR  (Preliminary Analytical Report)")
        self.rb_par.setObjectName('reportTypeRadio')
        self.report_type_group.addButton(self.rb_par, 1)
        rt_layout.addWidget(self.rb_par)

        rt_layout.addStretch()
        info_layout.addRow('Report Type:', report_type_widget)

        # Update Generate button label when report type changes
        self.report_type_group.buttonToggled.connect(self._update_generate_btn_label)

        # Fetch details button
        self.fetch_details_btn = QPushButton("🔍 Auto-Fetch Details from BS File")
        self.fetch_details_btn.setObjectName('fetchDetailsBtn')
        self.fetch_details_btn.clicked.connect(self._fetch_details_from_file)
        info_layout.addRow('', self.fetch_details_btn)
        
        # Round Off
        self.round_off = QCheckBox("Enable Rounding")
        self.round_off.setChecked(True)
        info_layout.addRow('Round Off:', self.round_off)
        
        # Rounding unit
        self.unit_widget = QWidget()
        unit_layout = QHBoxLayout(self.unit_widget)
        unit_layout.setContentsMargins(0, 0, 0, 0)
        self.unit_group = QButtonGroup()
        
        units = ['Thousands', 'Lakhs', 'Millions']
        for i, unit in enumerate(units):
            rb = QRadioButton(unit)
            self.unit_group.addButton(rb, i)
            unit_layout.addWidget(rb)
            if unit == 'Lakhs':
                rb.setChecked(True)
        
        unit_layout.addStretch()
        info_layout.addRow('Rounding Unit:', self.unit_widget)
        
        # Connect checkbox to enable/disable units
        self.round_off.toggled.connect(self.unit_widget.setEnabled)
        
        # Notes required
        self.notes_required = QCheckBox('Include Notes to Accounts analysis')
        self.notes_required.setChecked(True)
        info_layout.addRow('Notes Required:', self.notes_required)
        
        main_layout.addWidget(info_group)
        

        # ── File Upload ──
        files_group = QGroupBox('Financial Statement')
        files_layout = QVBoxLayout(files_group)
        files_layout.setSpacing(12)
        files_layout.setContentsMargins(16, 20, 16, 16)
        
        self.fs_upload = FileUploadPanel('Financial Statement File (BS, P&L, Notes)')
        self.fs_upload.fileSelected.connect(self._auto_fill_client_name)
        files_layout.addWidget(self.fs_upload)
        
        main_layout.addWidget(files_group)
        
        # ── Generate Button ──
        btn_row = QHBoxLayout()
        btn_row.addStretch()
        
        self.generate_btn = QPushButton('🚀  Generate FAR Workpaper')  # label updated dynamically
        self.generate_btn.setObjectName('generateFarBtn')
        self.generate_btn.setFixedHeight(48)
        self.generate_btn.setMinimumWidth(280)
        self.generate_btn.setFont(QFont('Segoe UI', 14, QFont.Weight.Bold))
        self.generate_btn.clicked.connect(self._on_generate)
        btn_row.addWidget(self.generate_btn)
        
        btn_row.addStretch()
        main_layout.addLayout(btn_row)
        
        # Validation error label
        self.error_label = QLabel('')
        self.error_label.setObjectName('errorLabel')
        self.error_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        main_layout.addWidget(self.error_label)
        
        main_layout.addStretch()
        
        scroll.setWidget(content)
        
        # Wrap scroll in main layout
        wrapper = QVBoxLayout(self)
        wrapper.setContentsMargins(0, 0, 0, 0)
        wrapper.addWidget(scroll)
        


    def _fetch_details_from_file(self):
        """Fetch client name and financial year from the selected Balance Sheet file."""
        if not self.fs_upload.is_loaded():
            QMessageBox.warning(self, 'No File Selected', 'Please select a Financial Statement file first.')
            return
            
        # Check date range validation before auto-fetching
        py_from = self._parse_date(self.py_from_date.text())
        py_to = self._parse_date(self.py_to_date.text())
        cy_from = self._parse_date(self.cy_from_date.text())
        cy_to = self._parse_date(self.cy_to_date.text())
        
        if not py_from or not py_to or not cy_from or not cy_to:
            QMessageBox.warning(self, 'Invalid Date Range', 'Please enter valid dates in DD-MM-YYYY format before auto-fetching.')
            return
            
        if py_from > py_to or cy_from > cy_to:
            QMessageBox.warning(self, 'Invalid Date Range', 'From Date cannot be greater than To Date. Please correct the date range before auto-fetching.')
            return
        
        filepath = self.fs_upload.get_paths()[0]
        try:
            from openpyxl import load_workbook
            from src.core.excel_parser import _extract_client_name, _find_date_columns
            
            wb = load_workbook(filepath, data_only=True, read_only=True)
            
            # Find BS sheet
            ws = None
            for name in ['BS', 'Balance Sheet', 'BalanceSheet']:
                if name in wb.sheetnames:
                    ws = wb[name]
                    break
            if ws is None:
                ws = wb.worksheets[0]
                
            client_name = _extract_client_name(ws)
            _, _, cy_year, _, _ = _find_date_columns(ws)
            wb.close()
            
            extracted_info = []
            if client_name:
                self.client_name.setText(client_name)
                extracted_info.append(f"Client Name: {client_name}")
            
            if cy_year:
                # Set CY
                cy_fy_text = f"{cy_year - 1}-{str(cy_year)[-2:]}"
                idx_cy = self.cy_year_sel.findText(cy_fy_text)
                if idx_cy >= 0:
                    self.cy_year_sel.setCurrentIndex(idx_cy)
                else:
                    self.cy_year_sel.addItem(cy_fy_text)
                    self.cy_year_sel.setCurrentText(cy_fy_text)
                self.cy_from_date.setText(f"01-04-{cy_year - 1}")
                self.cy_to_date.setText(f"31-03-{cy_year}")
                
                # Set PY
                py_fy_text = f"{cy_year - 2}-{str(cy_year - 1)[-2:]}"
                idx_py = self.py_year_sel.findText(py_fy_text)
                if idx_py >= 0:
                    self.py_year_sel.setCurrentIndex(idx_py)
                else:
                    self.py_year_sel.addItem(py_fy_text)
                    self.py_year_sel.setCurrentText(py_fy_text)
                self.py_from_date.setText(f"01-04-{cy_year - 2}")
                self.py_to_date.setText(f"31-03-{cy_year - 1}")
                
                extracted_info.append(f"CY: {cy_fy_text} | PY: {py_fy_text}")
                
            if extracted_info:
                QMessageBox.information(
                    self, 'Details Fetched',
                    "Successfully fetched details from Balance Sheet:\n\n" + "\n".join(extracted_info)
                )
            else:
                QMessageBox.warning(
                    self, 'Details Not Found',
                    "Could not extract Client Name or Financial Year from the selected Balance Sheet file."
                )
        except Exception as e:
            QMessageBox.critical(
                self, 'Error Fetching Details',
                f"Failed to parse details from the Balance Sheet file:\n\n{str(e)}"
            )

    def _parse_date(self, text):
        from PyQt6.QtCore import QDate
        import datetime
        try:
            dt = datetime.datetime.strptime(text.strip(), "%d-%m-%Y")
            return QDate(dt.year, dt.month, dt.day)
        except Exception:
            return None

    def _update_metrics(self):
        py_from = self._parse_date(self.py_from_date.text())
        py_to = self._parse_date(self.py_to_date.text())
        cy_from = self._parse_date(self.cy_from_date.text())
        cy_to = self._parse_date(self.cy_to_date.text())
        
        py_days = 0
        cy_days = 0
        
        # Calculate PY metrics
        if py_from and py_to:
            py_days, py_months, py_coverage = self._compute_metrics(py_from, py_to)
            self.py_metrics_label.setText(f"Days Covered: {py_days} | Months Covered: {py_months} | Coverage: {py_coverage}%")
        else:
            self.py_metrics_label.setText("Days Covered: — | Months Covered: — | Coverage: —")
            
        # Calculate CY metrics
        if cy_from and cy_to:
            cy_days, cy_months, cy_coverage = self._compute_metrics(cy_from, cy_to)
            self.cy_metrics_label.setText(f"Days Covered: {cy_days} | Months Covered: {cy_months} | Coverage: {cy_coverage}%")
        else:
            self.cy_metrics_label.setText("Days Covered: — | Months Covered: — | Coverage: —")
            
        # Warning visibility
        if py_from and py_to and cy_from and cy_to:
            periods_unequal = (py_days != cy_days)
            show_warning = periods_unequal and self.auto_annualize_cb.isChecked()
            self.warning_label.setVisible(show_warning)
        else:
            self.warning_label.setVisible(False)

    def _compute_metrics(self, from_qd, to_qd):
        from_dt = from_qd.toPyDate()
        to_dt = to_qd.toPyDate()
        days = (to_dt - from_dt).days + 1
        if days <= 0:
            return 0, 0.0, 0.0
        months = days / 30.4375
        if days >= 365:
            months = 12.0
        coverage = min(100.0, (days / 365.0) * 100.0)
        return days, round(months, 1), round(coverage, 1)

    def _get_financial_year_string(self):
        py_from = self._parse_date(self.py_from_date.text())
        py_to = self._parse_date(self.py_to_date.text())
        cy_from = self._parse_date(self.cy_from_date.text())
        cy_to = self._parse_date(self.cy_to_date.text())
        
        if cy_from and cy_to and py_from and py_to:
            return f"{self.cy_year_sel.currentText()} vs {self.py_year_sel.currentText()}"
        return "Custom Period"

    def _validate(self):
        """Validate all form fields."""
        errors = []
        
        if not self.client_name.text().strip():
            errors.append('Client Name is required')
        if not self.fs_upload.is_loaded():
            errors.append('Financial Statement file is required')
            
        py_from = self._parse_date(self.py_from_date.text())
        py_to = self._parse_date(self.py_to_date.text())
        cy_from = self._parse_date(self.cy_from_date.text())
        cy_to = self._parse_date(self.cy_to_date.text())
        
        if not py_from:
            errors.append('PY From Date must be in DD-MM-YYYY format')
        if not py_to:
            errors.append('PY To Date must be in DD-MM-YYYY format')
        if not cy_from:
            errors.append('CY From Date must be in DD-MM-YYYY format')
        if not cy_to:
            errors.append('CY To Date must be in DD-MM-YYYY format')
            
        if py_from and py_to and py_from > py_to:
            errors.append('PY From Date cannot be greater than PY To Date')
        if cy_from and cy_to and cy_from > cy_to:
            errors.append('CY From Date cannot be greater than CY To Date')
        
        if errors:
            self.error_label.setText('⚠️ ' + ' | '.join(errors))
            return False
        
        self.error_label.setText('')
        return True
    
    def _auto_fill_client_name(self, filepath):
        """
        Automatically populate the Client Name field from the selected BS workbook.
        Scans the first 5 rows of columns A-C to find the company name.
        Only fills if the field is currently empty (does not overwrite user input).
        """
        if self.client_name.text().strip():
            return  # Don't overwrite if user already typed something
        try:
            from openpyxl import load_workbook as _load_wb
            wb = _load_wb(filepath, data_only=True, read_only=True)
            ws = wb.worksheets[0]
            # Try to find 'BS' sheet
            for name in ['BS', 'Balance Sheet', 'BalanceSheet']:
                if name in wb.sheetnames:
                    ws = wb[name]
                    break
            extracted = ''
            for row_idx in range(1, 6):
                for col_idx in range(1, 4):
                    cell_val = ws.cell(row=row_idx, column=col_idx).value
                    if cell_val and isinstance(cell_val, str) and len(cell_val.strip()) > 3:
                        extracted = cell_val.strip()
                        break
                if extracted:
                    break
            wb.close()
            if extracted:
                self.client_name.setText(extracted)
                import logging
                logging.getLogger(__name__).info(
                    "Auto-filled client name from workbook: %s", extracted)
        except Exception as e:
            import logging
            logging.getLogger(__name__).warning(
                "Could not auto-fill client name from workbook: %s", e)

    def _update_generate_btn_label(self):
        """Update the Generate button label to reflect the selected report type."""
        rtype = 'FAR' if self.rb_far.isChecked() else 'PAR'
        self.generate_btn.setText(f'🚀  Generate {rtype} Workpaper')

    def _get_report_type(self):
        """Return the currently selected report type string."""
        return 'FAR' if self.rb_far.isChecked() else 'PAR'

    def _on_generate(self):
        """Handle Generate button click."""
        if not self._validate():
            return
            
        py_from = self._parse_date(self.py_from_date.text())
        py_to = self._parse_date(self.py_to_date.text())
        cy_from = self._parse_date(self.cy_from_date.text())
        cy_to = self._parse_date(self.cy_to_date.text())
        
        unit_btn = self.unit_group.checkedButton()
        unit_text = unit_btn.text() if unit_btn else 'Lakhs'
        
        fs_path = self.fs_upload.get_paths()[0] if self.fs_upload.is_loaded() else ''
        form_data = {
            'firm_name': self.firm_name.text().strip() or 'Audit Firm',
            'client_name': self.client_name.text().strip(),
            'financial_year': self._get_financial_year_string(),
            'cy_from_date': cy_from.toString('yyyy-MM-dd') if cy_from else '',
            'cy_to_date': cy_to.toString('yyyy-MM-dd') if cy_to else '',
            'py_from_date': py_from.toString('yyyy-MM-dd') if py_from else '',
            'py_to_date': py_to.toString('yyyy-MM-dd') if py_to else '',
            'auto_annualize': self.auto_annualize_cb.isChecked(),
            'round_off': self.round_off.isChecked(),
            'rounding_unit': unit_text,
            'notes_required': self.notes_required.isChecked(),
            'bs_file': fs_path,
            'pl_file': fs_path,
            'notes_files': [fs_path] if fs_path else [],
            'report_type': self._get_report_type(),
        }
        
        self.generateRequested.emit(form_data)
    
    def get_form_data(self):
        """Get current form data without triggering generation."""
        py_from = self._parse_date(self.py_from_date.text())
        py_to = self._parse_date(self.py_to_date.text())
        cy_from = self._parse_date(self.cy_from_date.text())
        cy_to = self._parse_date(self.cy_to_date.text())
        
        unit_btn = self.unit_group.checkedButton()
        unit_text = unit_btn.text() if unit_btn else 'Lakhs'
        
        fs_path = self.fs_upload.get_paths()[0] if self.fs_upload.is_loaded() else ''
        return {
            'firm_name': self.firm_name.text().strip() or 'Audit Firm',
            'client_name': self.client_name.text().strip(),
            'financial_year': self._get_financial_year_string(),
            'cy_from_date': cy_from.toString('yyyy-MM-dd') if cy_from else '',
            'cy_to_date': cy_to.toString('yyyy-MM-dd') if cy_to else '',
            'py_from_date': py_from.toString('yyyy-MM-dd') if py_from else '',
            'py_to_date': py_to.toString('yyyy-MM-dd') if py_to else '',
            'auto_annualize': self.auto_annualize_cb.isChecked(),
            'round_off': self.round_off.isChecked(),
            'rounding_unit': unit_text,
            'notes_required': self.notes_required.isChecked(),
            'bs_file': fs_path,
            'pl_file': fs_path,
            'notes_files': [fs_path] if fs_path else [],
            'report_type': self._get_report_type(),
        }
